import importlib
import os
import warnings
import pandas as pd # type: ignore
import openpyxl # type: ignore
import Utilities
import SystemDeskEnums

# Suppress specific warnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ProjectManager:
    def __init__(self):
        self.input_excel_file = None
        self.workbook = None
        self.project_info = None
        self.swc_info = None
        self.ib_data = None
        self.ports = None
        self.adt_primitive = None
        self.adt_composite = None
        self.SdApplication = Utilities.ConnectToSystemDesk()  # Establish connection to SystemDesk

    def read_user_defined_excel(self):
        """Prompts the user for an Excel file path, loads the workbook,
        and assigns each worksheet to a corresponding variable for easy access.
        """
        self.input_excel_file = input("Please provide the path of the input Excel file (e.g., E:\\sysdesk\\automation): ")
        self.workbook = openpyxl.load_workbook(self.input_excel_file, data_only=True)
        
        # Define worksheet names for better readability
        worksheet_names = [
            "Project_Info",
            "SWC_Info",
            "IB_data",
            "Ports",
            "ADT_Primitive",
            "ADT_Composite"
        ]
        
        # Assign each worksheet to a corresponding variable
        self.project_info, self.swc_info, self.ib_data, self.ports, self.adt_primitive, self.adt_composite = [
            self.workbook.worksheets[i] for i in range(len(worksheet_names))
        ]

    def read_columns(self, CurrentSheet, first_col, last_col, start_row=2, end_row=None):
        """Reads specified columns from the given sheet and returns the data as lists."""
        column_lists = []
        if end_row is None:
            end_row = CurrentSheet.max_row
        
        num_columns = ord(last_col) - ord(first_col) + 1
        valid_columns = [[] for _ in range(num_columns)]
        
        for row in range(start_row, end_row + 1):
            for col in range(num_columns):
                cell_value = CurrentSheet.cell(row=row, column=col + ord(first_col) - ord('A') + 1).value
                valid_columns[col].append(cell_value if cell_value is not None else None)
        
        # Filter out trailing None values from each column
        filtered_columns = []
        for column in valid_columns:
            while column and column[-1] is None:
                column.pop()
            filtered_columns.append(column)
        
        return filtered_columns

    def is_boolean(self, obj):
        """Checks if the object is of boolean type."""
        if "'bool'" in str(obj.__class__):
            return True
        elif "'list'" in str(obj.__class__) and bool(obj):
            return self.is_boolean(obj[0])
        elif "'tuple'" in str(obj.__class__) and bool(obj):
            return self.is_boolean(obj[0])
        else:
            return False

    def value_to_str(self, value):
        """Converts a value to its string representation."""
        if self.is_boolean(value):
            return "1" if value else "0"
        elif value == int(value):
            return str(int(value))
        else:
            return str(value)

    def project_config(self):
        """Configures project settings based on the project information."""
        self.Options = Utilities.Options
        sd_project = self.project_info['C4'].value
        self.Options.ProjectName = sd_project
        self.Options.OpenExistingProject = False
        self.Options.SystemName = "System"
        self.Options.UseFrontIndicatorEcus = True
        self.Options.CreateImplementations = True
        self.Options.ImportNetworkCommunicationElements = True
        self.Options.GenerateCode = True
        self.Options.License_SYD_MOD = True
        self.Options.BuildForVeos = True
        self.Options.SaveProject = True

    def create_project_and_packages(self):
        """Creates a new project and imports necessary packages."""
        Utilities.CreateProject()
        applRootDir = r"<InstallationDir>"
        templateFiles = [os.path.join(applRootDir, r"Templates\FolderStructure.arxml")]
        Utilities.ImportAutosarFilesAtProject(templateFiles)

    def cleanup_project_and_packages(self):
        """Cleans up the project by deleting specified packages."""
        rootPackage = self.SdApplication.ActiveProject.RootAutosar
        swComponentTypesPackage = rootPackage.ArPackages .Item("SwComponentTypes")
        
        # List of package names to be deleted
        for pkg_name in ["MySwc", "MyComposition"]:
            pkg = swComponentTypesPackage.ArPackages.Item(pkg_name)
            if pkg is not None:
                pkg.Delete()

    def create_swcs(self):
        """Creates software components based on the information in swc_info."""
        value_B2 = self.swc_info['B2'].value
        switcher = {
            'ApplicationSwComponentType': self.my_application_function,
            'ComplexDeviceDriverSwComponentType': self.my_complex_device_driver_function,
            'EcuAbstractionSwComponentType': self.my_ecu_abstraction_function,
            'NvBlockSwComponentType': self.my_nv_block_function,
            'ParameterSwComponentType': self.my_parameter_function,
            'SensorActuatorSwComponentType': self.my_sensor_actuator_function,
            'ServiceProxySwComponentType': self.my_service_proxy_function,
            'ServiceSwComponentType': self.my_service_function
        }
        
        func = switcher.get(value_B2, self.my_application_function)
        func()

    def my_application_function(self):
        """Handles the creation of application software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentApplSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("ApplSWC")
        CurrentSWC = CurrentApplSWCPkg.Elements.AddNewApplicationSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreateOtherIBElements(CurrentInternalBehaviors)
        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()
        # self.createrteevent(CurrentInternalBehaviors, RTEEventName, RunnableName)

    def my_complex_device_driver_function(self):
        """Handles the creation of complex device driver software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentCddSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("CddSWC")
        CurrentSWC = CurrentCddSWCPkg.Elements.AddNewComplexDeviceDriverSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_ecu_abstraction_function(self):
        """Handles the creation of ECU abstraction software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentEcuAbSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("EcuAbSWC")
        CurrentSWC = CurrentEcuAbSWCPkg.Elements.AddNewEcuAbstractionSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_nv_block_function(self):
        """Handles the creation of NV block software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentNvBlockSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("NvBlockSWC")
        CurrentSWC = CurrentNvBlockSWCPkg.Elements.AddNewNvBlockSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_parameter_function(self):
        """Handles the creation of parameter software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentPrmSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("PrmSWC")
        CurrentSWC = CurrentPrmSWCPkg.Elements.AddNewParameterSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value

        # Create ports and shared interfaces
        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_sensor_actuator_function(self):
        """Handles the creation of sensor actuator software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentSnsrActSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SnsrActSWC")
        CurrentSWC = CurrentSnsrActSWCPkg.Elements.AddNewSensorActuatorSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_service_proxy_function(self):
        """Handles the creation of service proxy software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentSrvcPrxySWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcPrxySWC")
        CurrentSWC = CurrentSrvcPrxySWCPkg.Elements.AddNewServiceProxySwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")
        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def my_service_function(self):
        """Handles the creation of service software components."""
        CurrentswComponentTypesPackage = self.SdApplication.ActiveProject.RootAutosar.ArPackages.Item("SwComponentTypes")
        CurrentSrvcSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcSWC")
        CurrentSWC = CurrentSrvcSWCPkg.Elements.AddNewServiceSwComponentType(self.swc_info['C2'].value)
        CurrentSWC.Symbol = self.swc_info['D2'].value
        CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(self.swc_info['E2'].value)

        # Handle termination and restart support
        termination_support = self.swc_info['F2'].value
        if termination_support == 'noSupport':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
        elif termination_support == 'canBeTerminated':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
        elif termination_support == 'canBeTerminatedAndRestarted':
            CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
        else:
            raise ValueError(f"Unexpected value for handleTerminationAndRestart: {termination_support}")

        CurrentInternalBehaviors.SupportsMultipleInstantiation = self.swc_info['G2'].value
        RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = self.read_columns(self.swc_info, 'H', 'M')

        for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
            CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
            CurrentRunnable.MinimumStartInterval = 0.0
            CurrentRunnable.CanBeInvokedConcurrently = c
            CurrentRunnable.Symbol = b
            self.createrteevent(CurrentInternalBehaviors,CurrentRunnable, d, e, f)

        self.CreatePorts(CurrentSWC)
        self.createsharedinterfaces()

    def CreatePorts(self,CurrentSWC):
        """Creates ports for the current software component."""
        PortType, PortName = self.read_columns(self.ports, 'B', 'C')
        
        for port_name, port_type in zip(PortName, PortType):
            if port_type == 'ReceiverPort':
                CurrentRport = CurrentSWC.Ports.AddNewRPortPrototype(port_name)
            elif port_type == 'SenderPort':
                CurrentPport = CurrentSWC.Ports.AddNewPPortPrototype(port_name)
            else:
                print(f"Unknown port type: {port_type} for port {port_name}")

    def createsharedinterfaces(self):
        """Creates shared interfaces based on the ports data."""
        IF_type_col, IF_name_col, DE_col, Argument_col = self.read_columns(self.ports, 'D', 'G')
        
        interface_collections = {
            'SenderReceiverInterface': [],
            'NvDataInterface': [],
            'ParameterInterface': [],
            'ModeSwitchInterface': [],
            'ClientServerInterface': [],
            'TriggerInterface': []
        }
        
        current_interface = None
        current_type = None
        
        for i in range(len(IF_name_col)):
            name = str(IF_name_col[i]).strip()
            interface_type = str(IF_type_col[i]).strip()
            data_element = str(DE_col[i]).strip()
            
            if name and name != 'nan':
                if current_interface:
                    if current_type in interface_collections:
                        interface_collections[current_type].append(current_interface)
                
                current_interface = {
                    'name': name,
                    'data_elements': [data_element] if data_element and data_element != 'nan' else []
                }
                current_type = interface_type  # Update the current type to the new interface type
            else:
                if current_interface and data_element and data_element != 'nan':
                    current_interface['data_elements'].append(data_element)
        
        if current_interface and current_type:
            if current_type in interface_collections:
                interface_collections[current_type].append(current_interface)
        
        interface_creation_map = {
            'SenderReceiverInterface': self.SenderReceiverInterface,
            'NvDataInterface': self.NvDataInterface,
            'ParameterInterface': self.ParameterInterface,
            'ModeSwitchInterface': self.ModeSwitchInterface,
            'ClientServerInterface': self.ClientServerInterface,
            'TriggerInterface': self.TriggerInterface
        }
        
        for interface_type, interfaces in interface_collections.items():
            if interface_type in interface_creation_map:
                creation_func = interface_creation_map[interface_type]
                for interface in interfaces:
                    try:
                        creation_func(interface['name'], interface['data_elements'])
                    except Exception as e:
                        print(f"Error creating {interface_type} {interface['name']}: {e}")

    def SenderReceiverInterface(self, currentIF_name, DataElements):
        """Creates a SenderReceiver interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/SenderReceiver")
        currentSRIF = SharedIFpkg.Elements.AddNewSenderReceiverInterface(currentIF_name)
        currentSRIF.SetNewIsService().SetValue(False)
        
        for itsDE in DataElements:
            currentSRIFDE = currentSRIF.DataElements.AddNew(itsDE)
            Utilities.SetDescription(currentSRIFDE, "Turn switch sensor position [left, off, right]")
            Utilities.SetInvalidationPolicy(currentSRIF, currentSRIFDE, SystemDeskEnums.HandleInvalidEnum.DontInvalidate)

    def NvDataInterface(self, currentIF_name, NvDatas):
        """Creates an NvData interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/NvData")
        currentNVIF = SharedIFpkg.Elements.AddNewNvDataInterface(currentIF_name)
        currentNVIF.SetNewIsService().SetValue(False)
        
        for itsDE in NvDatas:
            currentNVIFDE = currentNVIF.NvDatas.AddNew(itsDE)
            Utilities.SetDescription(currentNVIFDE, "Turn switch sensor position [left, off, right]")

    def ParameterInterface(self, currentIF_name, Parameters):
        """Creates a Parameter interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Parameter")
        currentPRMIF = SharedIFpkg.Elements.AddNewParameterInterface(currentIF_name)
        currentPRMIF.SetNewIsService().SetValue(False)
        
        for itsDE in Parameters:
            currentPRMIFDE = currentPRMIF.Parameters.AddNew(itsDE)
            Utilities.SetDescription(currentPRMIFDE, "Turn switch sensor position [left, off, right]")

    def ModeSwitchInterface(self, currentIF_name, ModeDeclaration):
        """Creates a ModeSwitch interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ModeSwitch")
        # Functionality for ModeSwitchInterface is not implemented yet

    def ClientServerInterface(self, currentIF_name, Operations):
        """Creates a ClientServer interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ClientServer")
        currentCSIF = SharedIFpkg.Elements.AddNewClientServerInterface(currentIF_name)
        currentCSIF.SetNewIsService().SetValue(False)
        
        for itsDE in Operations:
            currentCSIFDE = currentCSIF.Operations.AddNew(itsDE)
            Utilities.SetDescription(currentCSIFDE, "Turn switch sensor position [left, off, right]")

    def TriggerInterface(self, currentIF_name, trigger):
        """Creates a Trigger interface."""
        SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Trigger")
        # Functionality for TriggerInterface is not implemented yet

    def create_shared_parameter(self, CurrentInternalBehaviors, b):
        """Creates a shared parameter."""
        CurrentSharedParameter = CurrentInternalBehaviors.SharedParameters.AddNew(b)
        CurrentSharedParameter.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentSharedParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentSharedParameter, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentSharedParameter, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentSharedParameter)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_per_instance_parameter(self, CurrentInternalBehaviors, b):
        """Creates a per-instance parameter."""
        CurrentPerInstanceParameter = CurrentInternalBehaviors.PerInstanceParameters.AddNew(b)
        CurrentPerInstanceParameter.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentPerInstanceParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentPerInstanceParameter, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentPerInstanceParameter, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentPerInstanceParameter)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_implicit_inter_runnable_variable(self, CurrentInternalBehaviors, b):
        """Creates an implicit inter-runnable variable."""
        CurrentImplicitInterRunnableVariables = CurrentInternalBehaviors.ImplicitInterRunnableVariables.AddNew(b)
        CurrentImplicitInterRunnableVariables.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentImplicitInterRunnableVariables, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentImplicitInterRunnableVariables)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_explicit_inter_runnable_variable(self, CurrentInternalBehaviors, b):
        """Creates an explicit inter-runnable variable."""
        CurrentExplicitInterRunnableVariable = CurrentInternalBehaviors.ExplicitInterRunnableVariables.AddNew(b)
        CurrentExplicitInterRunnableVariable.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentExplicitInterRunnableVariable, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentExplicitInterRunnableVariable)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_per_instance_memory(self, CurrentInternalBehaviors, b):
        """Creates a per-instance memory."""
        CurrentPerInstanceMemory = CurrentInternalBehaviors.PerInstanceMemorys.AddNew(b) 
        # No additional setup is specified for per-instance memory

    def create_ar_typed_per_instance_memory(self, CurrentInternalBehaviors, b):
        """Creates an AR typed per-instance memory."""
        CurrentArTypedPerInstanceMemory = CurrentInternalBehaviors.ArTypedPerInstanceMemorys.AddNew(b)
        CurrentArTypedPerInstanceMemory.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentArTypedPerInstanceMemory, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentArTypedPerInstanceMemory)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_static_memory(self, CurrentInternalBehaviors, b):
        """Creates static memory."""
        CurrentStaticMemorys = CurrentInternalBehaviors.StaticMemorys.AddNew(b)
        CurrentStaticMemorys.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentStaticMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentStaticMemorys, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentStaticMemorys, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentStaticMemorys)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def create_constant_memory(self, CurrentInternalBehaviors, b):
        """Creates constant memory."""
        CurrentConstantMemorys = CurrentInternalBehaviors.ConstantMemorys.AddNew(b)
        CurrentConstantMemorys.Category = "VALUE"
        Utilities.SetSwImplPolicy(CurrentConstantMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
        Utilities.SetSwCalibrationAccess(CurrentConstantMemorys, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
        maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentConstantMemorys, "adt_max_count")
        maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
        Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
        maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
        Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
        maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentConstantMemorys)
        Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

    def CreateOtherIBElements(self, CurrentInternalBehaviors):
        """Creates other internal behavior elements based on the IB data."""
        IBVariableType, IBVariableName, ApplicationDataTypeName, Initvalue, AccessingRunnable, SwCalibrationAccess, SwImplementationPolicy = self.read_columns(self.ib_data, 'B', 'H')

        switcher = {
            'SharedParameter': self.create_shared_parameter,
            'PerInstanceParameter': self.create_per_instance_parameter,
            'ImplicitInterRunnableVariables': self.create_implicit_inter_runnable_variable,
            'ExplicitInterRunnableVariable': self.create_explicit_inter_runnable_variable,
            'PerInstanceMemory': self.create_per_instance_memory,
            'ArTypedPerInstanceMemory': self.create_ar_typed_per_instance_memory,
            'StaticMemory': self.create_static_memory,
            'ConstantMemory': self.create_constant_memory
        }

        for a, b in zip(IBVariableType, IBVariableName):
            func = switcher.get(a)
            if func:
                func(CurrentInternalBehaviors, b)
            else:
                print("No Other Internal Behaviour Elements present for this software component")

    def createrteevent(self, CurrentInternalBehaviors,RunnableName, RTEEventName, RTEEventType, RTEEventInfo):
        """Creates RTE events based on the event type."""
        if RTEEventType == 'AsynchronousServerCallReturnsEvent':
            self.AsynchronousServerCallReturnsEvent(CurrentInternalBehaviors,RTEEventName, RunnableName)
        elif RTEEventType == 'InitEvent':
            self.InitEvent(CurrentInternalBehaviors,RTEEventName, RunnableName)
        elif RTEEventType == 'BackgroundEvent':
            self.BackgroundEvent(CurrentInternalBehaviors,RTEEventName, RunnableName)
        elif RTEEventType == 'TimingEvent':
            self.TimingEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'DataReceiveErrorEvent':
            self.DataReceiveErrorEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'DataSendCompletedEvent':
            self.DataSendCompletedEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'DataWriteCompletedEvent':
            self.DataWriteCompletedEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'ExternalTriggerOccurredEvent':
            self.ExternalTriggerOccurredEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'InternalTriggerOccurredEvent':
            self.InternalTriggerOccurredEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'ModeSwitchedAckEvent':
            self.ModeSwitchedAckEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'OperationInvokedEvent':
            self.OperationInvokedEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'SwcModeManagerErrorEvent':
            self.SwcModeManagerErrorEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'DataReceivedEvent':
            self.DataReceivedEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'SwcModeSwitchEvent':
            self.SwcModeSwitchEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        elif RTEEventType == 'TransformerHardErrorEvent':
            self.TransformerHardErrorEvent(CurrentInternalBehaviors,RTEEventName, RunnableName, RTEEventInfo)
        else:
            print(f"Unrecognized event type: {RTEEventName}")

    def InitEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable):
        """Creates an initialization event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewInitEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable

    def TimingEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a timing event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewTimingEvent(CurrentRteEvent)
        CurrentEvent.Period = RTEEventInfo
        CurrentEvent.StartOnEventRef = AssociatedRunnable

    def AsynchronousServerCallReturnsEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates an asynchronous server call returns event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewAsynchronousServerCallReturnsEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def DataReceiveErrorEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a data receive error event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceiveErrorEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.DataIref = RTEEventInfo

    def DataReceivedEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a data received event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceivedEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.DataIref = RTEEventInfo

    def DataSendCompletedEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a data send completed event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataSendCompletedEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def DataWriteCompletedEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a data write completed event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataWriteCompletedEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def ExternalTriggerOccurredEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates an external trigger occurred event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewExternalTriggerOccurredEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.TriggerIref = RTEEventInfo

    def InternalTriggerOccurredEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates an internal trigger occurred event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewInternalTriggerOccurredEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.TriggerIref = RTEEventInfo

    def ModeSwitchedAckEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a mode switched acknowledgment event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewModeSwitchedAckEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def OperationInvokedEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates an operation invoked event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewOperationInvokedEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def SwcModeManagerErrorEvent(self,CurrentInternalBehaviors, CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a software component mode manager error event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeManagerErrorEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def DataReceivedEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a data received event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceivedEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.DataIref = RTEEventInfo

    def SwcModeSwitchEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a software component mode switch event."""
        CurrentEvent = self.CurrentInternalBehaviors.Events.AddNewSwcModeSwitchEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def TransformerHardErrorEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a transformer hard error event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewTransformerHardErrorEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def SwcModeSwitchEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a software component mode switch event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeSwitchEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def TransformerHardErrorEvent(self,CurrentInternalBehaviors, CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
        """Creates a transformer hard error event."""
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewTransformerHardErrorEvent(CurrentRteEvent)
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        CurrentEvent.EventSourceRef = RTEEventInfo

    def BackgroundEvent(self, CurrentInternalBehaviors,CurrentRteEvent, AssociatedRunnable):
        # Create a new background event in the current internal behaviors
        CurrentEvent = CurrentInternalBehaviors.Events.AddNewBackgroundEvent(CurrentRteEvent)
        # Set the reference to the associated runnable for this event
        CurrentEvent.StartOnEventRef = AssociatedRunnable
        




    def createcompumethod(self):
        """Creates computation methods based on the ADT_Primitive data source."""
        CompuMethodName, CompuMethodCategory, CompuScaleOROffset, EnumStatesORLSB, Unit = self.read_columns(self.adt_primitive, 'D', 'H')
        
        CompuMethodsPackage = Utilities.GetElementByPath("/SharedElements/CompuMethods")
        commonPackage = Utilities.GetElementByPath("/AUTOSAR/AUTOSAR_PhysicalUnits/Units")
        
        current_compu_method_name = None
        current_compu_method_category = None
        current_unit = None
        compu_scale = []  # List to hold scale or offset values
        enum_states = []  # List to hold enumeration states or LSB values
        
        for i in range(len(CompuMethodName)):
            if CompuMethodName[i]:
                if current_compu_method_name is not None:
                    currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
                    CurrentUnit = commonPackage.Elements.Item(current_unit)
                    
                    switcher = {
                        'IDENTICAL': self.handle_identical,
                        'TEXTTABLE': self.handle_texttable,
                        'LINEAR': self.handle_linear,
                        'SCALE_LINEAR': self.handle_scale_linear,
                        'SCALE_LINEAR_AND_TEXTTABLE': self.handle_scale_linear_and_texttable,
                        'RAT_FUNC': self.handle_rat_func,
                        'SCALE_RAT_FUNC': self.handle_scale_rat_func,
                        'SCALE_RATIONAL_AND_TEXTTABLE': self.handle_scale_rational_and_texttable,
                        'TAB_NOINTP': self.handle_tab_nointp,
                        'BITFIELD_TEXTTABLE': self.handle_bitfield_texttable,
                    }
                    
                    func = switcher.get(current_compu_method_category, self.handle_identical)
                    func(currentcompumethod, compu_scale, enum_states, CurrentUnit)

                current_compu_method_name = CompuMethodName[i]
                current_compu_method_category = CompuMethodCategory[i]
                current_unit = Unit[i]
                compu_scale = []  # Reset the scale list for the new method
                enum_states = []  # Reset the enum states list for the new method

            if i < len(CompuScaleOROffset) and CompuScaleOROffset[i] is not None:
                compu_scale.append(CompuScaleOROffset[i])

            if i < len(EnumStatesORLSB) and EnumStatesORLSB[i] is not None:
                enum_states.append(EnumStatesORLSB[i])

        if current_compu_method_name is not None:
            currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
            CurrentUnit = commonPackage.Elements.Item(current_unit)
            func = switcher.get(current_compu_method_category, self.handle_identical)
            func(currentcompumethod, compu_scale, enum_states, CurrentUnit)

    def handle_identical(self, currentcompumethod, NA, NA1, Current_Unit):
        """Handles identical computation methods."""
        # Implementation for identical computation methods
        pass  # Add your logic here

    def handle_texttable(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles text table computation methods."""
        # Implementation for text table computation methods
        pass  # Add your logic here

    def handle_linear(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles linear computation methods."""
        # Implementation for linear computation methods
        pass  # Add your logic here

    def handle_scale_linear(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles scale linear computation methods."""
        # Implementation for scale linear computation methods
        pass  # Add your logic here

    def handle_scale_linear_and_texttable(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles scale linear and text table computation methods."""
            # Implementation for scale linear and text table computation methods
        pass  # Add your logic here

    def handle_rat_func(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles rational function computation methods."""
        # Implementation for rational function computation methods
        pass  # Add your logic here

    def handle_scale_rat_func(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles scale rational function computation methods."""
        # Implementation for scale rational function computation methods
        pass  # Add your logic here

    def handle_scale_rational_and_texttable(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles scale rational and text table computation methods."""
        # Implementation for scale rational and text table computation methods
        pass  # Add your logic here

    def handle_tab_nointp(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles table without interpolation computation methods."""
        # Implementation for table without interpolation computation methods
        pass  # Add your logic here

    def handle_bitfield_texttable(self, currentcompumethod, scale, enum_states, Current_Unit):
        """Handles bitfield text table computation methods."""
        # Implementation for bitfield text table computation methods
        pass  # Add your logic here


    def createDC(self): 
        # Read columns from the adt_primitive data source 
        DataConstraintName, DataConstraintType, Min, Max = self.read_columns(self.adt_primitive, 'I', 'L') 

        # Get the Data Constraints package from the shared elements 
        DCPackage = Utilities.GetElementByPath("/SharedElements/DataConstr") 

        # Iterate through the data constraints 
        for a, b, c, d in zip(DataConstraintName, DataConstraintType, Min, Max): 
            constrLevel = 0  # Initialize constraint level 
            lowerLimit = c   # Set lower limit from the read data 
            upperLimit = d   # Set upper limit from the read data 

            # Check if the constraint type is 'internalConstrs' 
            if b == 'internalConstrs': 
                # Create a new internal data constraint 
                dataConstr = DCPackage.Elements.AddNewDataConstr(a) 
                dataConstrRule = dataConstr.DataConstrRules.AddNew() 

                # Set the constraint level if it's not None 
                if constrLevel is not None: 
                    dataConstrRule.ConstrLevel = constrLevel 

                # Set new internal constraints 
                internalConstrs = dataConstrRule.SetNewInternalConstrs() 

                # Set the lower limit if it is not None 
                if lowerLimit is not None: 
                    lowLimit = internalConstrs.SetNewLowerLimit() 
                    lowLimit.MixedContent.AddStringContent(self.value_to_str(lowerLimit)) 
                    lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed 

                # Set the upper limit if it is not None 
                if upperLimit is not None: 
                    upLimit = internalConstrs.SetNewUpperLimit() 
                    upLimit.MixedContent.AddStringContent(self.value_to_str(upperLimit)) 
                    upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed 

            # Check if the constraint type is 'physConstrs' 
            elif b == 'physConstrs': 
                # Create a new physical data constraint 
                dataConstr = DCPackage.Elements.AddNewDataConstr(a) 
                dataConstrRule = dataConstr.DataConstrRules.AddNew() 

                # Set the constraint level if it's not None 
                if constrLevel is not None: 
                    dataConstrRule.ConstrLevel = constrLevel 

                # Set new physical constraints 
                physConstrs = dataConstrRule.SetNewPhysConstrs() 

                # Set the lower limit if it is not None 
                if lowerLimit is not None: 
                    lowLimit = physConstrs.SetNewLowerLimit() 
                    lowLimit.MixedContent.AddStringContent(self.value_to_str(lowerLimit)) 
                    lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed 

                # Set the upper limit if it is not None 
                if upperLimit is not None: 
                    upLimit = physConstrs.SetNewUpperLimit() 
                    upLimit.MixedContent.AddStringContent(self.value_to_str(upperLimit)) 
                    upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed 

            # If the constraint type is neither 'internalConstrs' nor 'physConstrs' 
            else: 
                if b is not None: 
                    print(f"Value of b: {b}")  # Print the value of b for debugging purposes 

    def main(self):  # Change 'Main' to 'main'
        # Execute the main sequence of functions for project setup 
        self.read_user_defined_excel()  # Read user-defined data from an Excel file 
        self.project_config()         # Configure the project settings 
        self.create_project_and_packages()  # Create the project and its packages 
        self.cleanup_project_and_packages()  # Clean up any temporary project data 
        self.create_swcs()           # Create software components 
        self.createcompumethod()    # Create computation methods 
        self.createDC()             # Create data constraints 

# Entry point of the script 
if __name__ == "__main__": 
    project_manager = ProjectManager()  # Create an instance of ProjectManager 
    project_manager.main()  # Call the main method
