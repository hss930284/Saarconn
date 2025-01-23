##############################################################################################################################################################################################################################
####################################################################### Importing Libraries ############################################################################################################################################
###################################################################################################################################################################################################################################################################
from ast import Global
import importlib
import os
import warnings
from pathlib import Path

from Tutorial import CreateInterRunnableVariables

# Suppress warnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Third-party library imports
import pandas as pd # type: ignore
import openpyxl # type: ignore

# Local module imports
import Utilities
import SystemDeskEnums  # type: ignore

# Reloading modules
importlib.reload(Utilities)
importlib.reload(SystemDeskEnums)

# Connecting to SystemDesk (get COM object)
SdApplication = Utilities.ConnectToSystemDesk()


##############################################################################################################################################################################################################################
####################################################################### Excel-related functions ############################################################################################################################################
###################################################################################################################################################################################################################################################################

def ReadUserDefinedExcel():
    """
    Reads the user-provided Excel file and loads the required data into global variables.
    Additionally, reads the value of cell C4 from the first sheet.
    """
    global Input_Excel_File, xls, workbook

    Input_Excel_File = input("Please provide the path of input Excel file (e.g., E:\\sysdesk\\automation): ")
    workbook = openpyxl.load_workbook(Input_Excel_File, data_only=True)
    xls = pd.ExcelFile(Input_Excel_File)



def read_columns(CurrentSheet, first_col, last_col):
    # Initialize a list to hold the column lists
    column_lists = []

    # Determine the last row with any value in the specified columns
    max_row = 2  # Start from row 2
    while True:
        # Check if there's any value in the specified columns for the current row
        has_value = any(CurrentSheet.cell(row=max_row, column=col).value is not None for col in range(ord(first_col) - ord('A') + 1, ord(last_col) - ord('A') + 2))
        if not has_value:
            break
        max_row += 1  # Move to the next row

    # Now read values from row 2 to the last valid row
    for row in range(2, max_row):
        column_data = []
        for col in range(ord(first_col) - ord('A') + 1, ord(last_col) - ord('A') + 2):
            column_data.append(CurrentSheet.cell(row=row, column=col).value)
        column_lists.append(column_data)

    # Unzip the list of lists into separate lists for each column
    column_lists = list(zip(*column_lists))

    # Return the individual column lists
    return [list(column) for column in column_lists]


##############################################################################################################################################################################################################################
####################################################################### Project Configuration Functions ############################################################################################################################################
###################################################################################################################################################################################################################################################################

def ProjectConfig():
    """
    Configures the project with predefined settings.
    """
    global Options
    Options = Utilities.Options

    # Read the first sheet
    first_sheet = workbook.worksheets[0]
    
    # Get the value of cell C4
    sd_project = first_sheet['C4'].value

    Options.ProjectName = sd_project
    Options.OpenExistingProject = False
    Options.SystemName = "System"
    Options.UseFrontIndicatorEcus = True
    Options.CreateImplementations = True
    Options.ImportNetworkCommunicationElements = True
    Options.GenerateCode = True
    Options.License_SYD_MOD = True
    Options.BuildForVeos = True
    Options.SaveProject = True

def CreateProjectAndPackages():
    """
    Creates a new project and imports predefined packages.
    """
    Utilities.CreateProject()
    applRootDir = r"<InstallationDir>"
    templateFiles = [os.path.join(applRootDir, r"Templates\FolderStructure.arxml")]
    Utilities.ImportAutosarFilesAtProject(templateFiles)

def CleanupProjectAndPackages():
    """
    Removes unnecessary elements imported from template files.
    """
    rootPackage = SdApplication.ActiveProject.RootAutosar
    swComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")

    # Delete unnecessary packages
    for pkg_name in ["MySwc", "MyComposition"]:
        pkg = swComponentTypesPackage.ArPackages.Item(pkg_name)
        if pkg is not None:
            pkg.Delete()



##############################################################################################################################################################################################################################
####################################################################### SWC Creation Functions ############################################################################################################################################
###################################################################################################################################################################################################################################################################

def CreateSwcs():

    global rootPackage, value_B2, SWC_Info

    # Read the SWC_Info sheet
    SWC_Info = workbook.worksheets[1]
    
    # Get the value of cell B2
    value_B2 = SWC_Info['B2'].value

    rootPackage = SdApplication.ActiveProject.RootAutosar

    if value_B2 == 'ApplicationSwComponentType':
        ApplicationSwComponentType()
    elif value_B2 == 'ComplexDeviceDriverSwComponentType':
        ComplexDeviceDriverSwComponentType()
    elif value_B2 == 'EcuAbstractionSwComponentType':
        EcuAbstractionSwComponentType()
    elif value_B2 == 'NvBlockSwComponentType':
        NvBlockSwComponentType()
    elif value_B2 == 'ParameterSwComponentType':
        ParameterSwComponentType()
    elif value_B2 == 'SensorActuatorSwComponentType':
        SensorActuatorSwComponentType()
    elif value_B2 == 'ServiceProxySwComponentType':
        ServiceProxySwComponentType()
    elif value_B2 == 'ServiceSwComponentType':
        ServiceSwComponentType()
    else:
        print("In sheet SWC_Info, Please provide correct software component type")

def ApplicationSwComponentType():
    global CurrentswComponentTypesPackage, CurrentApplSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentApplSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("ApplSWC")
    CurrentSWC = CurrentApplSWCPkg.Elements.AddNewApplicationSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)

    CreateOtherIBElements(CurrentInternalBehaviors)
    
#     CreatePorts()
#     createsharedinterfaces()

def ComplexDeviceDriverSwComponentType():

    global CurrentswComponentTypesPackage, CurrentCddSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentCddSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("CddSWC")
    CurrentSWC = CurrentCddSWCPkg.Elements.AddNewComplexDeviceDriverSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()

def EcuAbstractionSwComponentType():
    global CurrentswComponentTypesPackage, CurrentEcuAbSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentEcuAbSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("EcuAbSWC")
    CurrentSWC = CurrentEcuAbSWCPkg.Elements.AddNewEcuAbstractionSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()

def NvBlockSwComponentType():
    global CurrentswComponentTypesPackage, CurrentNvBlockSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentNvBlockSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("NvBlockSWC")
    CurrentSWC = CurrentNvBlockSWCPkg.Elements.AddNewNvBlockSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()


def ParameterSwComponentType():
    global CurrentswComponentTypesPackage, CurrentPrmSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentPrmSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("PrmSWC")
    CurrentSWC = CurrentPrmSWCPkg.Elements.AddNewParameterSwComponentType(SWC_Info['C2'].value)  #software component name
    
#     CreatePorts()
#     createsharedinterfaces()


def SensorActuatorSwComponentType():
    global CurrentswComponentTypesPackage, CurrentSnsrActSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSnsrActSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SnsrActSWC")
    CurrentSWC = CurrentSnsrActSWCPkg.Elements.AddNewSensorActuatorSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()

def ServiceProxySwComponentType():
    global CurrentswComponentTypesPackage, CurrentSrvcPrxySWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSrvcPrxySWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcPrxySWC")
    CurrentSWC = CurrentSrvcPrxySWCPkg.Elements.AddNewServiceProxySwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()

def ServiceSwComponentType():
    global CurrentswComponentTypesPackage, CurrentSrvcSWCPkg, CurrentSWC, CurrentInternalBehaviors

    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSrvcSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcSWC")
    CurrentSWC = CurrentSrvcSWCPkg.Elements.AddNewServiceSwComponentType(SWC_Info['C2'].value)  #software component name
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)   #Internal Behavior

    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value

         # Call the function with the desired columns
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')

    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    
#     CreatePorts()
#     createsharedinterfaces()


##############################################################################################################################################################################################################################
########################################## RTE Event creation related definitions ############################################################################################################################################
###################################################################################################################################################################################################################################################################

def CreateOtherIBElements(CurrentInternalBehaviors):

    global IB_data

    IB_data = workbook.worksheets[2]

    IBVariableType,IBVariableName,ApplicationDataTypeName,Initvalue,AccessingRunnable,SwCalibrationAccess,SwImplementationPolicy= read_columns(IB_data, 'B', 'H')

    for a,b,c,d,e,f,g in zip(IBVariableType,IBVariableName,ApplicationDataTypeName,Initvalue,AccessingRunnable,SwCalibrationAccess,SwImplementationPolicy):

        if a == 'SharedParameter':
            CurrentSharedParameter = CurrentInternalBehaviors.SharedParameters.AddNew(b)

            CurrentSharedParameter.Category = "VALUE"
            # Utilities.SetDescription(CurrentSharedParameter, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentSharedParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentSharedParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentSharedParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentSharedParameter, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentSharedParameter)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)


        elif a == 'PerInstanceParameter':
            CurrentPerInstanceParameter = CurrentInternalBehaviors.PerInstanceParameters.AddNew(b)

            CurrentPerInstanceParameter.Category = "VALUE"
            # Utilities.SetDescription(CurrentPerInstanceParameter, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentPerInstanceParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentPerInstanceParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentPerInstanceParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentPerInstanceParameter, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentPerInstanceParameter)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
            
        elif a == 'ImplicitInterRunnableVariables':
            CurrentImplicitInterRunnableVariables = CurrentInternalBehaviors.ImplicitInterRunnableVariables.AddNew(b)

            CurrentImplicitInterRunnableVariables.Category = "VALUE"
            # Utilities.SetDescription(CurrentImplicitInterRunnableVariables, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentImplicitInterRunnableVariables,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentImplicitInterRunnableVariables,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentImplicitInterRunnableVariables, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentImplicitInterRunnableVariables)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

           

        elif a == 'ExplicitInterRunnableVariable':
            CurrentExplicitInterRunnableVariable = CurrentInternalBehaviors.ExplicitInterRunnableVariables.AddNew(b)

            CurrentExplicitInterRunnableVariable.Category = "VALUE"
            # Utilities.SetDescription(CurrentExplicitInterRunnableVariable, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentExplicitInterRunnableVariable,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentExplicitInterRunnableVariable,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentExplicitInterRunnableVariable, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentExplicitInterRunnableVariable)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)



        elif a == 'PerInstanceMemory':
            CurrentPerInstanceMemory = CurrentInternalBehaviors.PerInstanceMemorys.AddNew(b)

            # CurrentPerInstanceMemory.Category = "VALUE"
            # # Utilities.SetDescription(CurrentPerInstanceMemory, "Determines the duration of the on/off state of the indicator lights.")
            # Utilities.SetSwImplPolicy(CurrentPerInstanceMemory, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # # Utilities.SetSwCalibrationAccess(CurrentPerInstanceMemory,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            # Utilities.SetSwCalibrationAccess(CurrentPerInstanceMemory,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            # maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentPerInstanceMemory, "adt_max_count")
            # # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            # Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            # maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            # Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            # maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentPerInstanceMemory)
            # Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)




        elif a == 'ArTypedPerInstanceMemory':
            CurrentArTypedPerInstanceMemory = CurrentInternalBehaviors.ArTypedPerInstanceMemorys.AddNew(b)

            CurrentArTypedPerInstanceMemory.Category = "VALUE"
            # Utilities.SetDescription(CurrentArTypedPerInstanceMemory, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentArTypedPerInstanceMemory,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentArTypedPerInstanceMemory,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentArTypedPerInstanceMemory, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentArTypedPerInstanceMemory)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)


        elif a == 'StaticMemory':
            CurrentStaticMemorys = CurrentInternalBehaviors.StaticMemorys.AddNew(b)

            CurrentStaticMemorys.Category = "VALUE"
            # Utilities.SetDescription(CurrentStaticMemorys, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentStaticMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentStaticMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentStaticMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentStaticMemorys, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentStaticMemorys)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

        elif a == 'ConstantMemory':
            CurrentConstantMemorys = CurrentInternalBehaviors.ConstantMemorys.AddNew(b)

            CurrentConstantMemorys.Category = "VALUE"
            # Utilities.SetDescription(CurrentConstantMemorys, "Determines the duration of the on/off state of the indicator lights.")
            Utilities.SetSwImplPolicy(CurrentConstantMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
            # Utilities.SetSwCalibrationAccess(CurrentConstantMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite,displayFormat="%3d")
            Utilities.SetSwCalibrationAccess(CurrentConstantMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentConstantMemorys, "adt_max_count")
            # maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "CM_IDENTICAL")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentConstantMemorys)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)



        else:
            print("In sheet IB_data, Please provide correct IB Element or create manually in the tool")

        


    




##############################################################################################################################################################################################################################
########################################## RTE Event creation related definitions ############################################################################################################################################
###################################################################################################################################################################################################################################################################


def createrteevent(RunnableName,RTEEventName,RTEEventType,RTEEventInfo):
    
    if RTEEventType == 'AsynchronousServerCallReturnsEvent':
        AsynchronousServerCallReturnsEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'InitEvent':
        InitEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'BackgroundEvent':
        BackgroundEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'TimingEvent':
        TimingEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'AsynchronousServerCallReturnsEvent':
        AsynchronousServerCallReturnsEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataReceiveErrorEvent':
        DataReceiveErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataSendCompletedEvent':
        DataSendCompletedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataWriteCompletedEvent':
        DataWriteCompletedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'ExternalTriggerOccurredEvent':
        ExternalTriggerOccurredEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'InternalTriggerOccurredEvent':
        InternalTriggerOccurredEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'ModeSwitchedAckEvent':
        ModeSwitchedAckEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'OperationInvokedEvent':
        OperationInvokedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'SwcModeManagerErrorEvent':
        SwcModeManagerErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataReceivedEvent':
        DataReceivedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'SwcModeSwitchEvent':
        SwcModeSwitchEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'TransformerHardErrorEvent':
        TransformerHardErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    else:
        print(f"Unrecognized event type: {RTEEventName}")
    
def InitEvent(CurrentRteEvent,AssociatedRunnable):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInitEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable

def TimingEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTimingEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.Period = RTEEventInfo
    CurrentEvent.StartOnEventRef = AssociatedRunnable

def AsynchronousServerCallReturnsEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewAsynchronousServerCallReturnsEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def DataReceiveErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceiveErrorEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo

def DataReceivedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceivedEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo

def DataSendCompletedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataSendCompletedEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def DataWriteCompletedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataWriteCompletedEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def ExternalTriggerOccurredEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewExternalTriggerOccurredEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.TriggerIref = RTEEventInfo

def InternalTriggerOccurredEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInternalTriggerOccurredEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def ModeSwitchedAckEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewModeSwitchedAckEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def OperationInvokedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewOperationInvokedEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.OperationIref = RTEEventInfo

def SwcModeManagerErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeManagerErrorEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.ModeGroupIref = RTEEventInfo

def SwcModeSwitchEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeSwitchEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.ModeIref = RTEEventInfo #exited and entered
    CurrentEvent.Activation = RTEEventInfo

def TransformerHardErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTransformerHardErrorEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo

def BackgroundEvent(CurrentRteEvent,AssociatedRunnable):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewBackgroundEvent(CurrentRteEvent)
    # Utilities.SetDescription(tssEvent, "TimingEvent for TssPreprocessing Runnable [10 ms period].")
    CurrentEvent.StartOnEventRef = AssociatedRunnable



##############################################################################################################################################################################################################################
####################################################################### Main Function ############################################################################################################################################
###################################################################################################################################################################################################################################################################


def Main():
    ReadUserDefinedExcel()
    ProjectConfig()
    CreateProjectAndPackages()
    CleanupProjectAndPackages()
    CreateSwcs()

if __name__ == "__main__":
    Main()
    print("Ready")