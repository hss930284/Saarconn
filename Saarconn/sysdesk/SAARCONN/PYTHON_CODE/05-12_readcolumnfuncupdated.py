from ast import Global
import importlib
import os
import warnings
from pathlib import Path
from Tutorial import CreateInterRunnableVariables
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import pandas as pd # type: ignore
import openpyxl # type: ignore
import Utilities
import SystemDeskEnums 
importlib.reload(Utilities)
importlib.reload(SystemDeskEnums)
SdApplication = Utilities.ConnectToSystemDesk()

def ReadUserDefinedExcel():
    global Input_Excel_File, xls, workbook,Project_Info,SWC_Info,IB_data,Ports, ADT_Primitive, ADT_Composite
    Input_Excel_File = input("Please provide the path of input Excel file (e.g., E:\\sysdesk\\automation): ")
    workbook = openpyxl.load_workbook(Input_Excel_File, data_only=True)
    xls = pd.ExcelFile(Input_Excel_File)
    Project_Info = workbook.worksheets[0]
    SWC_Info = workbook.worksheets[1]
    IB_data = workbook.worksheets[2]
    Ports = workbook.worksheets[3]
    ADT_Primitive = workbook.worksheets[4]
    ADT_Composite = workbook.worksheets[5]
def read_columns(CurrentSheet, first_col, last_col, start_row=2, end_row=None):
    column_lists = []
    if end_row is None:
        end_row = CurrentSheet.max_row  
    num_columns = ord(last_col) - ord(first_col) + 1
    valid_columns = [[] for _ in range(num_columns)]
    for row in range(start_row, end_row + 1):
        for col in range(num_columns):
            cell_value = CurrentSheet.cell(row=row, column=col + ord(first_col) - ord('A') + 1).value 
            if cell_value is not None:
                valid_columns[col].append(cell_value)
            else:
                valid_columns[col].append(None)
    filtered_columns = []
    for column in valid_columns:
        while column and column[-1] is None:
            column.pop()
        filtered_columns.append(column)
    return filtered_columns
def IsBoolean(obj):
    if "'bool'" in str(obj.__class__):
        return True
    elif "'list'" in str(obj.__class__) and bool(obj):
        return IsBoolean(obj[0])
    elif "'tuple'" in str(obj.__class__) and bool(obj):
        return IsBoolean(obj[0])
    else:
        return False
def Value2Str(value):
    if IsBoolean(value):
        if value:
            return "1"
        else:
            return "0"
    elif value == int(value):
        return str(int(value))
    else:
        return str(value)

def ProjectConfig():
    global Options
    Options = Utilities.Options
    sd_project = Project_Info['C4'].value
    Options.ProjectName = sd_project
    Options.OpenExistingProject = False
    Options.SystemName = "System"
    Options.UseFrontIndicatorEcus = True
    Options.CreateImplementations = True
    Options.ImportNetworkCommunicationElements = True
    Options.GenerateCode = True
    Options.License_SYD_MOD = True
    Options.BuildForVeos = True
    Options.SaveProject = True
def CreateProjectAndPackages():
    Utilities.CreateProject()
    applRootDir = r"<InstallationDir>"
    templateFiles = [os.path.join(applRootDir, r"Templates\FolderStructure.arxml")]
    Utilities.ImportAutosarFilesAtProject(templateFiles)
def CleanupProjectAndPackages():
    global rootPackage
    rootPackage = SdApplication.ActiveProject.RootAutosar
    swComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    for pkg_name in ["MySwc", "MyComposition"]:
        pkg = swComponentTypesPackage.ArPackages.Item(pkg_name)
        if pkg is not None:
            pkg.Delete()

def CreateSwcs():
    value_B2 = SWC_Info['B2'].value
    if value_B2 == 'ApplicationSwComponentType':
        ApplicationSwComponentType()
    elif value_B2 == 'ComplexDeviceDriverSwComponentType':
        ComplexDeviceDriverSwComponentType()
    elif value_B2 == 'EcuAbstractionSwComponentType':
        EcuAbstractionSwComponentType()
    elif value_B2 == 'NvBlockSwComponentType':
        NvBlockSwComponentType()
    elif value_B2 == 'ParameterSwComponentType':
        ParameterSwComponentType()
    elif value_B2 == 'SensorActuatorSwComponentType':
        SensorActuatorSwComponentType()
    elif value_B2 == 'ServiceProxySwComponentType':
        ServiceProxySwComponentType()
    elif value_B2 == 'ServiceSwComponentType':
        ServiceSwComponentType()
    else:
        print("In sheet SWC_Info, Please provide correct software component type")
def ApplicationSwComponentType():
    global CurrentswComponentTypesPackage, CurrentApplSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentApplSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("ApplSWC")
    CurrentSWC = CurrentApplSWCPkg.Elements.AddNewApplicationSwComponentType(SWC_Info['C2'].value) 
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value) 
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreateOtherIBElements(CurrentInternalBehaviors)
    CreatePorts()
    createsharedinterfaces()

def ComplexDeviceDriverSwComponentType():
    global CurrentswComponentTypesPackage, CurrentCddSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentCddSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("CddSWC")
    CurrentSWC = CurrentCddSWCPkg.Elements.AddNewComplexDeviceDriverSwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def EcuAbstractionSwComponentType():
    global CurrentswComponentTypesPackage, CurrentEcuAbSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentEcuAbSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("EcuAbSWC")
    CurrentSWC = CurrentEcuAbSWCPkg.Elements.AddNewEcuAbstractionSwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def NvBlockSwComponentType():
    global CurrentswComponentTypesPackage, CurrentNvBlockSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentNvBlockSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("NvBlockSWC")
    CurrentSWC = CurrentNvBlockSWCPkg.Elements.AddNewNvBlockSwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def ParameterSwComponentType():
    global CurrentswComponentTypesPackage, CurrentPrmSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentPrmSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("PrmSWC")
    CurrentSWC = CurrentPrmSWCPkg.Elements.AddNewParameterSwComponentType(SWC_Info['C2'].value)
    CreatePorts()
    createsharedinterfaces()

def SensorActuatorSwComponentType():
    global CurrentswComponentTypesPackage, CurrentSnsrActSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSnsrActSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SnsrActSWC")
    CurrentSWC = CurrentSnsrActSWCPkg.Elements.AddNewSensorActuatorSwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def ServiceProxySwComponentType():
    global CurrentswComponentTypesPackage, CurrentSrvcPrxySWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSrvcPrxySWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcPrxySWC")
    CurrentSWC = CurrentSrvcPrxySWCPkg.Elements.AddNewServiceProxySwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def ServiceSwComponentType():
    global CurrentswComponentTypesPackage, CurrentSrvcSWCPkg, CurrentSWC, CurrentInternalBehaviors
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    CurrentSrvcSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcSWC")
    CurrentSWC = CurrentSrvcSWCPkg.Elements.AddNewServiceSwComponentType(SWC_Info['C2'].value)
    CurrentSWC.Symbol = SWC_Info['D2'].value
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(SWC_Info['E2'].value)
    if SWC_Info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif SWC_Info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif SWC_Info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {SWC_Info['F2'].value}")
    CurrentInternalBehaviors.SupportsMultipleInstantiation = SWC_Info['G2'].value
    RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo= read_columns(SWC_Info, 'H', 'M')
    for a,b,c,d,e,f in zip(RunnableName,RunnableSymbol,canBeInvokedConcurrently,RTEEventName,RTEEventType,RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        CurrentRunnable.MinimumStartInterval = 0.0
        CurrentRunnable.CanBeInvokedConcurrently = c
        CurrentRunnable.Symbol = b
        createrteevent(CurrentRunnable,d,e,f)
    CreatePorts()
    createsharedinterfaces()

def CreatePorts():
    PortType,PortName= read_columns(Ports, 'B', 'C')
    for port_name, port_type in zip(PortName, PortType):
        if port_type == 'ReceiverPort':
            CurrentRport = CurrentSWC.Ports.AddNewRPortPrototype(port_name)
        elif port_type == 'SenderPort':
            CurrentPport = CurrentSWC.Ports.AddNewPPortPrototype(port_name)
        else:
            print(f"Unknown port type: {port_type} for port {port_name}")

def createsharedinterfaces():
    IF_type_col,IF_name_col, DE_col, Argument_col = read_columns(Ports, 'D', 'G')
    interface_collections = {
        'SenderReceiverInterface': [],
        'NvDataInterface': [],
        'ParameterInterface': [],
        'ModeSwitchInterface': [],
        'ClientServerInterface': [],
        'TriggerInterface': []}
    current_interface = None
    current_type = None
    for i in range(len(IF_name_col)):
        name = str(IF_name_col[i]).strip()
        interface_type = str(IF_type_col[i]).strip()
        data_element = str(DE_col[i]).strip()
        if name and name != 'nan':
            if current_interface:
                if current_type in interface_collections:
                    interface_collections[current_type].append(current_interface)
            current_interface = {
                'name': name,
                'data_elements': [data_element] if data_element and data_element != 'nan' else [] }
            current_type = interface_type
        else:
            if current_interface and data_element and data_element != 'nan':
                current_interface['data_elements'].append(data_element)
    if current_interface and current_type:
        if current_type in interface_collections:
            interface_collections[current_type].append(current_interface)
    interface_creation_map = {
        'SenderReceiverInterface': SenderReceiverInterface,
        'NvDataInterface': NvDataInterface,
        'ParameterInterface': ParameterInterface,
        'ModeSwitchInterface': ModeSwitchInterface,
        'ClientServerInterface': ClientServerInterface,
        'TriggerInterface': TriggerInterface   }
    for interface_type, interfaces in interface_collections.items():
        if interface_type in interface_creation_map:
            creation_func = interface_creation_map[interface_type]
            for interface in interfaces:
                try:
                    creation_func(interface['name'], interface['data_elements'])
                except Exception as e:
                    print(f"Error creating {interface_type} {interface['name']}: {e}")

def SenderReceiverInterface(currentIF_name, DataElements):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/SenderReceiver")
    currentSRIF = SharedIFpkg.Elements.AddNewSenderReceiverInterface(currentIF_name)
    currentSRIF.SetNewIsService().SetValue(False)
    for itsDE in DataElements:
        currentSRIFDE = currentSRIF.DataElements.AddNew(itsDE)
        Utilities.SetDescription(currentSRIFDE, "Turn switch sensor position [left, off, right]")
        Utilities.SetInvalidationPolicy(currentSRIF, currentSRIFDE, SystemDeskEnums.HandleInvalidEnum.DontInvalidate)
def NvDataInterface(currentIF_name, NvDatas):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/NvData")
    currentNVIF = SharedIFpkg.Elements.AddNewNvDataInterface(currentIF_name)
    currentNVIF.SetNewIsService().SetValue(False)
    for itsDE in NvDatas:
        currentNVIFDE = currentNVIF.NvDatas.AddNew(itsDE)
        Utilities.SetDescription(currentNVIFDE, "Turn switch sensor position [left, off, right]")
def ParameterInterface(currentIF_name, Parameters):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Parameter")
    currentPRMIF = SharedIFpkg.Elements.AddNewParameterInterface(currentIF_name)
    currentPRMIF.SetNewIsService().SetValue(False)
    for itsDE in Parameters:
        currentPRMIFDE = currentPRMIF.Parameters.AddNew(itsDE)
        Utilities.SetDescription(currentPRMIFDE, "Turn switch sensor position [left, off, right]")
def ModeSwitchInterface(currentIF_name, ModeDeclaration):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ModeSwitch")
def ClientServerInterface(currentIF_name, Operations):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ClientServer")
    currentCSIF = SharedIFpkg.Elements.AddNewClientServerInterface(currentIF_name)
    currentCSIF.SetNewIsService().SetValue(False)
    for itsDE in Operations:
        currentCSIFDE = currentCSIF.Operations.AddNew(itsDE)
        Utilities.SetDescription(currentCSIFDE, "Turn switch sensor position [left, off, right]") 
def TriggerInterface(currentIF_name, trigger):
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Trigger")

def CreateOtherIBElements(CurrentInternalBehaviors):
    IBVariableType,IBVariableName,ApplicationDataTypeName,Initvalue,AccessingRunnable,SwCalibrationAccess,SwImplementationPolicy= read_columns(IB_data, 'B', 'H')
    for a,b,c,d,e,f,g in zip(IBVariableType,IBVariableName,ApplicationDataTypeName,Initvalue,AccessingRunnable,SwCalibrationAccess,SwImplementationPolicy):
        if a == 'SharedParameter':
            CurrentSharedParameter = CurrentInternalBehaviors.SharedParameters.AddNew(b)
            CurrentSharedParameter.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentSharedParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentSharedParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentSharedParameter, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentSharedParameter)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'PerInstanceParameter':
            CurrentPerInstanceParameter = CurrentInternalBehaviors.PerInstanceParameters.AddNew(b)
            CurrentPerInstanceParameter.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentPerInstanceParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentPerInstanceParameter,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentPerInstanceParameter, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentPerInstanceParameter)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'ImplicitInterRunnableVariables':
            CurrentImplicitInterRunnableVariables = CurrentInternalBehaviors.ImplicitInterRunnableVariables.AddNew(b)
            CurrentImplicitInterRunnableVariables.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentImplicitInterRunnableVariables,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentImplicitInterRunnableVariables, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentImplicitInterRunnableVariables)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'ExplicitInterRunnableVariable':
            CurrentExplicitInterRunnableVariable = CurrentInternalBehaviors.ExplicitInterRunnableVariables.AddNew(b)
            CurrentExplicitInterRunnableVariable.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentExplicitInterRunnableVariable,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentExplicitInterRunnableVariable, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentExplicitInterRunnableVariable)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'PerInstanceMemory':
            CurrentPerInstanceMemory = CurrentInternalBehaviors.PerInstanceMemorys.AddNew(b)
        elif a == 'ArTypedPerInstanceMemory':
            CurrentArTypedPerInstanceMemory = CurrentInternalBehaviors.ArTypedPerInstanceMemorys.AddNew(b)
            CurrentArTypedPerInstanceMemory.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentArTypedPerInstanceMemory,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentArTypedPerInstanceMemory, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentArTypedPerInstanceMemory)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'StaticMemory':
            CurrentStaticMemorys = CurrentInternalBehaviors.StaticMemorys.AddNew(b)
            CurrentStaticMemorys.Category = "VALUE"            
            Utilities.SetSwImplPolicy(CurrentStaticMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentStaticMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentStaticMemorys, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentStaticMemorys)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        elif a == 'ConstantMemory':
            CurrentConstantMemorys = CurrentInternalBehaviors.ConstantMemorys.AddNew(b)
            CurrentConstantMemorys.Category = "VALUE"
            Utilities.SetSwImplPolicy(CurrentConstantMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
            Utilities.SetSwCalibrationAccess(CurrentConstantMemorys,SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
            maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentConstantMemorys, "adt_max_count")
            maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
            Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
            maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
            Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
            maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentConstantMemorys)
            Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)
        else:
            print("In sheet IB_data, Please provide correct IB Element or create manually in the tool")

def createrteevent(RunnableName,RTEEventName,RTEEventType,RTEEventInfo):
    if RTEEventType == 'AsynchronousServerCallReturnsEvent':
        AsynchronousServerCallReturnsEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'InitEvent':
        InitEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'BackgroundEvent':
        BackgroundEvent(RTEEventName,RunnableName)
    elif RTEEventType == 'TimingEvent':
        TimingEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'AsynchronousServerCallReturnsEvent':
        AsynchronousServerCallReturnsEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataReceiveErrorEvent':
        DataReceiveErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataSendCompletedEvent':
        DataSendCompletedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataWriteCompletedEvent':
        DataWriteCompletedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'ExternalTriggerOccurredEvent':
        ExternalTriggerOccurredEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'InternalTriggerOccurredEvent':
        InternalTriggerOccurredEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'ModeSwitchedAckEvent':
        ModeSwitchedAckEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'OperationInvokedEvent':
        OperationInvokedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'SwcModeManagerErrorEvent':
        SwcModeManagerErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'DataReceivedEvent':
        DataReceivedEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'SwcModeSwitchEvent':
        SwcModeSwitchEvent(RTEEventName,RunnableName,RTEEventInfo)
    elif RTEEventType == 'TransformerHardErrorEvent':
        TransformerHardErrorEvent(RTEEventName,RunnableName,RTEEventInfo)
    else:
        print(f"Unrecognized event type: {RTEEventName}")

def InitEvent(CurrentRteEvent,AssociatedRunnable):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInitEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
def TimingEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTimingEvent(CurrentRteEvent)
    CurrentEvent.Period = RTEEventInfo
    CurrentEvent.StartOnEventRef = AssociatedRunnable
def AsynchronousServerCallReturnsEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewAsynchronousServerCallReturnsEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo
def DataReceiveErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceiveErrorEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo
def DataReceivedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceivedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo
def DataSendCompletedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataSendCompletedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo
def DataWriteCompletedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataWriteCompletedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def ExternalTriggerOccurredEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewExternalTriggerOccurredEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.TriggerIref = RTEEventInfo
def InternalTriggerOccurredEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInternalTriggerOccurredEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo
def ModeSwitchedAckEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewModeSwitchedAckEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo
def OperationInvokedEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewOperationInvokedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.OperationIref = RTEEventInfo
def SwcModeManagerErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeManagerErrorEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.ModeGroupIref = RTEEventInfo
def SwcModeSwitchEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeSwitchEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.ModeIref = RTEEventInfo
    CurrentEvent.Activation = RTEEventInfo
def TransformerHardErrorEvent(CurrentRteEvent,AssociatedRunnable,RTEEventInfo):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTransformerHardErrorEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo
def BackgroundEvent(CurrentRteEvent,AssociatedRunnable):
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewBackgroundEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable

def createcompumethod():
    CompuMethodName, CompuMethodCategory, CompuScaleOROffset, EnumStatesORLSB, Unit = read_columns(ADT_Primitive, 'D', 'H')
    CompuMethodsPackage = Utilities.GetElementByPath("/SharedElements/CompuMethods")
    commonPackage = Utilities.GetElementByPath("/AUTOSAR/AUTOSAR_PhysicalUnits/Units")
    current_compu_method_name = None
    current_compu_method_category = None
    current_unit = None
    compu_scale = []
    enum_states = []
    for i in range(len(CompuMethodName)):
        if CompuMethodName[i]:
            if current_compu_method_name is not None:
                currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
                CurrentUnit = commonPackage.Elements.Item(current_unit)
                switcher = {
                    'IDENTICAL': handle_identical,
                    'TEXTTABLE': handle_texttable,
                    'LINEAR': handle_linear,
                    'SCALE_LINEAR': handle_scale_linear,
                    'SCALE_LINEAR_AND_TEXTTABLE': handle_scale_linear_and_texttable,
                    'RAT_FUNC': handle_rat_func,
                    'SCALE_RAT_FUNC': handle_scale_rat_func,
                    'SCALE_RATIONAL_AND_TEXTTABLE': handle_scale_rational_and_texttable,
                    'TAB_NOINTP': handle_tab_nointp,
                    'BITFIELD_TEXTTABLE': handle_bitfield_texttable,}
                func = switcher.get(current_compu_method_category, handle_identical)
                func(currentcompumethod, compu_scale, enum_states, CurrentUnit)
            current_compu_method_name = CompuMethodName[i]
            current_compu_method_category = CompuMethodCategory[i]
            current_unit = Unit[i]
            compu_scale = []
            enum_states = []
        if i < len(CompuScaleOROffset) and CompuScaleOROffset[i] is not None:
            compu_scale.append(CompuScaleOROffset[i])
        if i < len(EnumStatesORLSB) and EnumStatesORLSB[i] is not None:
            enum_states.append(EnumStatesORLSB[i])
    if current_compu_method_name is not None:
        currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
        CurrentUnit = commonPackage.Elements.Item(current_unit)
        switcher = {
            'IDENTICAL': handle_identical,
            'TEXTTABLE': handle_texttable,
            'LINEAR': handle_linear,
            'SCALE_LINEAR': handle_scale_linear,
            'SCALE_LINEAR_AND_TEXTTABLE': handle_scale_linear_and_texttable,
            'RAT_FUNC': handle_rat_func,
            'SCALE_RAT_FUNC': handle_scale_rat_func,
            'SCALE_RATIONAL_AND_TEXTTABLE': handle_scale_rational_and_texttable,
            'TAB_NOINTP': handle_tab_nointp,
            'BITFIELD_TEXTTABLE': handle_bitfield_texttable,}
        func = switcher.get(current_compu_method_category, handle_identical)
        func(currentcompumethod, compu_scale, enum_states, CurrentUnit)

def handle_identical(currentcompumethod,NA, NA1, Current_Unit):
    if NA is None or NA == "":
        print("CompuMethodInfo is not applicable for IDENTICAL category.")
    if NA1 is None or NA1 == "":
        print("CompuMethodInfo is not applicable for IDENTICAL category.")
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for IDENTICAL category.")
    currentcompumethod.Category = 'IDENTICAL'
    currentcompumethod.UnitRef = Current_Unit
def handle_texttable(currentcompumethod,Compu_Scale, Enum_States, Current_Unit):
    defaultValue=None
    if Compu_Scale is None or Compu_Scale == "":
        print("CompuMethodInfo is not applicable for TEXTTABLE category.")
    if Enum_States is None or Enum_States == "":
        print("CompuMethodInfo is not applicable for TEXTTABLE category.")
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for TEXTTABLE category.")    
    currentcompumethod.Category = 'TEXTTABLE'
    currentcompumethod.DisplayFormat = r"%g"
    currentcompumethod.UnitRef = Current_Unit
    compuInternalToPhys = currentcompumethod.SetNewCompuInternalToPhys()
    for a, b in zip(Compu_Scale, Enum_States):
        lowerLimit = a
        upperLimit = a
        symbol = b
        compuScale = compuInternalToPhys.CompuScales.AddNew()
        compuScale.Symbol = symbol
        compuScaleLL = compuScale.SetNewLowerLimit()
        compuScaleLL.MixedContent.AddStringContent(str(lowerLimit))
        compuScaleLL.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        compuScaleUL = compuScale.SetNewUpperLimit()
        compuScaleUL.MixedContent.AddStringContent(str(upperLimit))
        compuScaleUL.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        compuScale.ShortLabel = symbol
        compuConst = compuScale.SetNewCompuConst()
        compuConst.Vt = symbol
    if defaultValue is not None:
        print(defaultValue)
        compuInternalToPhys.CompuDefaultValue.Vt = defaultValue
def handle_linear(currentcompumethod,CompuScaleOROffset, EnumStatesORLSB, Current_Unit):
    if CompuScaleOROffset is None or CompuScaleOROffset == "":
        print("CompuMethodInfo is not applicable for LINEAR category.")
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for LINEAR category.")
def handle_scale_linear(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_LINEAR category.")
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_LINEAR category.")
def handle_scale_linear_and_texttable(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_LINEAR_AND_TEXTTABLE category.")
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_LINEAR_AND_TEXTTABLE category.")
def handle_rat_func(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for RAT_FUNC category.")
    if unit is None or unit == "":
        print("Unit is not applicable for RAT_FUNC category.")
def handle_scale_rat_func(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_RAT_FUNC category.")
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_RAT_FUNC category.")
def handle_scale_rational_and_texttable(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_RATIONAL_AND_TEXTTABLE category.")
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_RATIONAL_AND_TEXTTABLE category.")
def handle_tab_nointp(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for TAB_NOINTP category.")
    if unit is None or unit == "":
        print("Unit is not applicable for TAB_NOINTP category.")
def handle_bitfield_texttable(compu_method_info, unit):
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for BITFIELD_TEXTTABLE category.")
    if unit is None or unit == "":
        print("Unit is not applicable for BITFIELD_TEXTTABLE category.")

def createDC():
    DataConstraintName,	DataConstraintType,	Min, Max = read_columns(ADT_Primitive, 'I', 'L')
    DCPackage = Utilities.GetElementByPath("/SharedElements/DataConstr")
    for a, b, c, d in zip(DataConstraintName,	DataConstraintType,	Min, Max):
        constrLevel=0
        lowerLimit=c
        upperLimit=d
        if b == 'internalConstrs':
            dataConstr = DCPackage.Elements.AddNewDataConstr(a)
            dataConstrRule = dataConstr.DataConstrRules.AddNew()
            if constrLevel != None:
                dataConstrRule.ConstrLevel = constrLevel
            internalConstrs = dataConstrRule.SetNewInternalConstrs()
            if lowerLimit != None:
                lowLimit = internalConstrs.SetNewLowerLimit()
                lowLimit.MixedContent.AddStringContent(Value2Str(lowerLimit))
                lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
            if upperLimit != None:
                upLimit = internalConstrs.SetNewUpperLimit()
                upLimit.MixedContent.AddStringContent(Value2Str(upperLimit))
                upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        elif b == 'physConstrs':
            dataConstr = DCPackage.Elements.AddNewDataConstr(a)
            dataConstrRule = dataConstr.DataConstrRules.AddNew()
            if constrLevel != None:
                dataConstrRule.ConstrLevel = constrLevel
            physConstrs = dataConstrRule.SetNewPhysConstrs()
            if lowerLimit != None:
                lowLimit = physConstrs.SetNewLowerLimit()
                lowLimit.MixedContent.AddStringContent(Value2Str(lowerLimit))
                lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
            if upperLimit != None:
                upLimit = physConstrs.SetNewUpperLimit()
                upLimit.MixedContent.AddStringContent(Value2Str(upperLimit))
                upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        else:
            if b is not None:
                print(f"Value of b: {b}")    

def Main():
    ReadUserDefinedExcel()
    ProjectConfig()
    CreateProjectAndPackages()
    CleanupProjectAndPackages()
    CreateSwcs()
    createcompumethod()
    createDC()
if __name__ == "__main__":
    Main()
    print("Ready")