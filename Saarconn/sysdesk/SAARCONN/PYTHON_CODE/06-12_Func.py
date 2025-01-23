#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

# Import necessary modules and functions
from ast import Global  # Import Global from ast module (not used in the provided code)
import importlib  # Import importlib for dynamic module loading
import os  # Import os for operating system dependent functionality
import warnings  # Import warnings to manage warning messages
from pathlib import Path  # Import Path for filesystem path manipulations

# Import custom modules
from Tutorial import CreateInterRunnableVariables  # Import a specific function or class from the Tutorial module

# Suppress specific warnings from openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Import pandas for data manipulation and analysis
import pandas as pd  # type: ignore

# Import openpyxl for reading/writing Excel files
import openpyxl  # type: ignore

# Import custom utility functions
import Utilities  # Import the Utilities module
import SystemDeskEnums  # Import the SystemDeskEnums module

# Reload the Utilities and SystemDeskEnums modules to ensure the latest changes are applied
importlib.reload(Utilities)
importlib.reload(SystemDeskEnums)

# Establish a connection to the SystemDesk application using a utility function
SdApplication = Utilities.ConnectToSystemDesk()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def read_user_defined_excel():
    """
    Prompts the user for an Excel file path, loads the workbook,
    and assigns each worksheet to a corresponding variable for easy access.
    """
    global input_excel_file, xls, workbook, project_info, swc_info, ib_data, ports, adt_primitive, adt_composite
    
    # Prompt the user to input the path of the Excel file
    input_excel_file = input("Please provide the path of the input Excel file (e.g., E:\\sysdesk\\automation): ")
    
    # Load the workbook and create a pandas ExcelFile object
    workbook = openpyxl.load_workbook(input_excel_file, data_only=True)
    xls = pd.ExcelFile(input_excel_file)
    
    # Define worksheet names for better readability
    worksheet_names = [
        "Project_Info", 
        "SWC_Info", 
        "IB_data", 
        "Ports", 
        "ADT_Primitive", 
        "ADT_Composite"
    ]
    
    # Assign each worksheet to a corresponding variable
    project_info, swc_info, ib_data, ports, adt_primitive, adt_composite = [
        workbook.worksheets[i] for i in range(len(worksheet_names))
    ]

def read_columns(CurrentSheet, first_col, last_col, start_row=2, end_row=None):
    # Initialize a list to hold the column data
    column_lists = []
    
    # If end_row is not specified, set it to the maximum row of the current sheet
    if end_row is None:
        end_row = CurrentSheet.max_row  
    
    # Calculate the number of columns to read based on the provided first and last column letters
    num_columns = ord(last_col) - ord(first_col) + 1
    
    # Create a list of lists to hold valid column data
    valid_columns = [[] for _ in range(num_columns)]
    
    # Loop through each row in the specified range
    for row in range(start_row, end_row + 1):
        # Loop through each column in the specified range
        for col in range(num_columns):
            # Get the cell value from the current row and column
            cell_value = CurrentSheet.cell(row=row, column=col + ord(first_col) - ord('A') + 1).value 
            
            # If the cell value is not None, append it to the corresponding column list
            if cell_value is not None:
                valid_columns[col].append(cell_value)
            else:
                # If the cell is empty, append None to maintain the structure
                valid_columns[col].append(None)
    
    # Filter out trailing None values from each column
    filtered_columns = []
    for column in valid_columns:
        while column and column[-1] is None:
            column.pop()  # Remove None values from the end of the column
        filtered_columns.append(column)  # Add the cleaned column to the result list
    
    return filtered_columns  # Return the list of filtered columns

def IsBoolean(obj):
    # Check if the object is of boolean type
    if "'bool'" in str(obj.__class__):
        return True
    # If the object is a non-empty list or tuple, check the first element recursively
    elif "'list'" in str(obj.__class__) and bool(obj):
        return IsBoolean(obj[0])
    elif "'tuple'" in str(obj.__class__) and bool(obj):
        return IsBoolean(obj[0])
    else:
        return False  # Return False if the object is not a boolean

def Value2Str(value):
    # Convert a value to its string representation
    if IsBoolean(value):
        # If the value is boolean, return "1" for True and "0" for False
        if value:
            return "1"
        else:
            return "0"
    # If the value is an integer, convert it to string without decimal
    elif value == int(value):
        return str(int(value))
    else:
        # For all other types, convert the value to string
        return str(value)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def ProjectConfig():
    # Declare a global variable to hold project options
    global Options
    # Initialize Options with the Utilities.Options object
    Options = Utilities.Options
    
    # Retrieve the project name from the Project_Info dictionary
    sd_project = project_info['C4'].value
    Options.ProjectName = sd_project  # Set the project name in Options
    
    # Configure various project settings
    Options.OpenExistingProject = False  # Indicate that we are creating a new project
    Options.SystemName = "System"  # Set the system name
    Options.UseFrontIndicatorEcus = True  # Enable the use of front indicator ECUs
    Options.CreateImplementations = True  # Allow the creation of implementations
    Options.ImportNetworkCommunicationElements = True  # Enable import of network communication elements
    Options.GenerateCode = True  # Enable code generation
    Options.License_SYD_MOD = True  # Set licensing for SYD_MOD
    Options.BuildForVeos = True  # Indicate that the project should be built for Veos
    Options.SaveProject = True  # Enable saving of the project

def CreateProjectAndPackages():
    # Create a new project using the Utilities function
    Utilities.CreateProject()
    
    # Define the root directory for the application installation
    applRootDir = r"<InstallationDir>"
    
    # Specify the template files to be imported into the project
    templateFiles = [os.path.join(applRootDir, r"Templates\FolderStructure.arxml")]
    
    # Import the specified AUTOSAR files into the project
    Utilities.ImportAutosarFilesAtProject(templateFiles)

def CleanupProjectAndPackages():
    # Declare a global variable to hold the root package of the project
    global rootPackage
    # Get the root AUTOSAR package of the active project
    rootPackage = SdApplication.ActiveProject.RootAutosar
    
    # Access the SwComponentTypes package within the root package
    swComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # List of package names to be deleted
    for pkg_name in ["MySwc", "MyComposition"]:
        # Attempt to retrieve the package by name
        pkg = swComponentTypesPackage.ArPackages.Item(pkg_name)
        # If the package exists, delete it
        if pkg is not None:
            pkg.Delete()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def CreateSwcs():
    # Retrieve the value from the swc_info dictionary at key 'B2'
    value_B2 = swc_info['B2'].value
    
    # Define a switcher dictionary mapping component types to their corresponding functions
    switcher = {
        'ApplicationSwComponentType': my_application_function,  # Change the function name here
        'ComplexDeviceDriverSwComponentType': my_complex_device_driver_function,  # Change the function name here
        'EcuAbstractionSwComponentType': my_ecu_abstraction_function,  # Change the function name here
        'NvBlockSwComponentType': my_nv_block_function,  # Change the function name here
        'ParameterSwComponentType': my_parameter_function,  # Change the function name here
        'SensorActuatorSwComponentType': my_sensor_actuator_function,  # Change the function name here
        'ServiceProxySwComponentType': my_service_proxy_function,  # Change the function name here
        'ServiceSwComponentType': my_service_function  # Change the function name here
    }
    
    # Get the function from the switcher dictionary, defaulting to my_application_function
    func = switcher.get(value_B2, my_application_function)  # Change the function name here
    
    # Call the function
    func()

def my_application_function():
    # Declare global variables to be used within this function
    global CurrentswComponentTypesPackage, CurrentApplSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the package for software component types
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the application software component package
    CurrentApplSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("ApplSWC")
    
    # Create a new application software component using the value from swc_info at key 'C2'
    CurrentSWC = CurrentApplSWCPkg.Elements.AddNewApplicationSwComponentType(swc_info['C2'].value) 
    
    # Set the symbol for the current software component using the value from swc_info at key 'D2'
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component using the value from swc_info at key 'E2'
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value) 
    
    # Determine the termination and restart support based on the value from swc_info at key 'F2'
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if the value for handleTerminationAndRestart is unexpected
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set the support for multiple instantiation based on the value from swc_info at key 'G2'
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns H to M from swc_info to get runnable details
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the details of each runnable and create them in the internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)  # Add a new runnable
        CurrentRunnable.MinimumStartInterval = 0.0  # Set the minimum start interval
        CurrentRunnable.CanBeInvokedConcurrently = c  # Set concurrency support
        CurrentRunnable.Symbol = b  # Set the symbol for the runnable
        
        # Create the RTE event for the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create other internal behavior elements
    CreateOtherIBElements(CurrentInternalBehaviors)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_complex_device_driver_function():
    # Declare global variables to hold current component types and packages
    global CurrentswComponentTypesPackage, CurrentCddSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the 'SwComponentTypes' package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the 'CddSWC' package within the 'SwComponentTypes' package
    CurrentCddSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("CddSWC")
    
    # Create a new complex device driver software component type using information from swc_info
    CurrentSWC = CurrentCddSWCPkg.Elements.AddNewComplexDeviceDriverSwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current software component
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Determine the termination and restart support based on the value in swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set whether the internal behavior supports multiple instantiation
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read runnable properties from swc_info for columns H to M
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the properties of each runnable and add them to the internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        # Add a new runnable with the specified name
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        
        # Set the minimum start interval for the runnable
        CurrentRunnable.MinimumStartInterval = 0.0
        
        # Set whether the runnable can be invoked concurrently
        CurrentRunnable.CanBeInvokedConcurrently = c
        
        # Set the symbol for the runnable
        CurrentRunnable.Symbol = b
        
        # Create the RTE event for the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_ecu_abstraction_function():
    # Declare global variables to hold current component type packages and elements
    global CurrentswComponentTypesPackage, CurrentEcuAbSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "EcuAbSWC" package within the SwComponentTypes package
    CurrentEcuAbSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("EcuAbSWC")
    
    # Create a new ECU Abstraction Software Component (SWC) using the value from swc_info
    CurrentSWC = CurrentEcuAbSWCPkg.Elements.AddNewEcuAbstractionSwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current SWC
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add new internal behaviors to the current SWC
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Determine the termination and restart support based on the value in swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set the support for multiple instantiation based on swc_info
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns for runnable information from swc_info
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the runnable information and create runnables for the current internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        # Add a new runnable with the name from the current iteration
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        
        # Set the minimum start interval for the runnable
        CurrentRunnable.MinimumStartInterval = 0.0
        
        # Set whether the runnable can be invoked concurrently
        CurrentRunnable.CanBeInvokedConcurrently = c
        
        # Set the symbol for the current runnable
        CurrentRunnable.Symbol = b
        
        # Create the RTE event for the current runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current SWC
    CreatePorts()
    
    # Create shared interfaces for the current SWC
    createsharedinterfaces()

def my_nv_block_function():
    # Declare global variables to be used within the function
    global CurrentswComponentTypesPackage, CurrentNvBlockSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "NvBlockSWC" package within the "SwComponentTypes" package
    CurrentNvBlockSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("NvBlockSWC")
    
    # Create a new NvBlockSwComponentType using the value from swc_info
    CurrentSWC = CurrentNvBlockSWCPkg.Elements.AddNewNvBlockSwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current software component
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Set the termination and restart handling based on the value in swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set whether the internal behavior supports multiple instantiation
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns from swc_info for runnable properties
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the properties of each runnable and add them to the internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        # Create a new runnable with the specified name
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        
        # Set the minimum start interval for the runnable
        CurrentRunnable.MinimumStartInterval = 0.0
        
        # Set whether the runnable can be invoked concurrently
        CurrentRunnable.CanBeInvokedConcurrently = c
        
        # Set the symbol for the runnable
        CurrentRunnable.Symbol = b
        
        # Create the RTE event associated with the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_parameter_function():
    # Declare global variables to be used within the function
    global CurrentswComponentTypesPackage, CurrentPrmSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "PrmSWC" package within the "SwComponentTypes" package
    CurrentPrmSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("PrmSWC")
    
    # Add a new Parameter Software Component Type using the value from swc_info
    CurrentSWC = CurrentPrmSWCPkg.Elements.AddNewParameterSwComponentType(swc_info['C2'].value)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_sensor_actuator_function():
    # Declare global variables to be used within the function
    global CurrentswComponentTypesPackage, CurrentSnsrActSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "SnsrActSWC" package within the "SwComponentTypes" package
    CurrentSnsrActSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SnsrActSWC")
    
    # Add a new Sensor Actuator Software Component Type using the value from swc_info
    CurrentSWC = CurrentSnsrActSWCPkg.Elements.AddNewSensorActuatorSwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current software component using the value from swc_info
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component using the value from swc_info
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Determine the termination and restart support based on the value from swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set the support for multiple instantiation based on the value from swc_info
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns from swc_info for runnable properties
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the properties of each runnable and add them to the current internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        # Add a new runnable with the specified name
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        
        # Set the minimum start interval for the runnable
        CurrentRunnable.MinimumStartInterval = 0.0
        
        # Set whether the runnable can be invoked concurrently
        CurrentRunnable.CanBeInvokedConcurrently = c
        
        # Set the symbol for the runnable
        CurrentRunnable.Symbol = b
        
        # Create the RTE event for the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_service_proxy_function():
    # Declare global variables to be used within the function
    global CurrentswComponentTypesPackage, CurrentSrvcPrxySWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "SrvcPrxySWC" package within "SwComponentTypes"
    CurrentSrvcPrxySWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcPrxySWC")
    
    # Create a new Service Proxy Software Component Type using the value from swc_info
    CurrentSWC = CurrentSrvcPrxySWCPkg.Elements.AddNewServiceProxySwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current software component
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Set the handleTerminationAndRestart property based on the value in swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set the SupportsMultipleInstantiation property based on the value in swc_info
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns from swc_info for runnable properties
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the properties of each runnable and add them to the internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)  # Add a new runnable
        CurrentRunnable.MinimumStartInterval = 0.0  # Set the minimum start interval
        CurrentRunnable.CanBeInvokedConcurrently = c  # Set concurrency capability
        CurrentRunnable.Symbol = b  # Set the symbol for the runnable
        
        # Create RTE events for the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

def my_service_function():
    # Declare global variables to hold current component types and packages
    global CurrentswComponentTypesPackage, CurrentSrvcSWCPkg, CurrentSWC, CurrentInternalBehaviors
    
    # Access the "SwComponentTypes" package from the root package
    CurrentswComponentTypesPackage = rootPackage.ArPackages.Item("SwComponentTypes")
    
    # Access the "SrvcSWC" package within the "SwComponentTypes" package
    CurrentSrvcSWCPkg = CurrentswComponentTypesPackage.ArPackages.Item("SrvcSWC")
    
    # Add a new service software component type using the value from swc_info
    CurrentSWC = CurrentSrvcSWCPkg.Elements.AddNewServiceSwComponentType(swc_info['C2'].value)
    
    # Set the symbol for the current software component
    CurrentSWC.Symbol = swc_info['D2'].value
    
    # Add a new internal behavior to the current software component
    CurrentInternalBehaviors = CurrentSWC.InternalBehaviors.AddNew(swc_info['E2'].value)
    
    # Determine the termination and restart handling based on the value in swc_info
    if swc_info['F2'].value == 'noSupport':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.NoSupport
    elif swc_info['F2'].value == 'canBeTerminated':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminated
    elif swc_info['F2'].value == 'canBeTerminatedAndRestarted':
        CurrentInternalBehaviors.handleTerminationAndRestart = SystemDeskEnums.HandleTerminationAndRestartEnum.CanBeTerminatedAndRestarted
    else:
        # Raise an error if an unexpected value is encountered
        raise ValueError(f"Unexpected value for handleTerminationAndRestart: {swc_info['F2'].value}")
    
    # Set whether the internal behavior supports multiple instantiation
    CurrentInternalBehaviors.SupportsMultipleInstantiation = swc_info['G2'].value
    
    # Read columns from swc_info for runnable properties
    RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo = read_columns(swc_info, 'H', 'M')
    
    # Loop through the properties of each runnable and add them to the internal behaviors
    for a, b, c, d, e, f in zip(RunnableName, RunnableSymbol, canBeInvokedConcurrently, RTEEventName, RTEEventType, RTEEventInfo):
        # Add a new runnable with the specified name
        CurrentRunnable = CurrentInternalBehaviors.Runnables.AddNew(a)
        
        # Set the minimum start interval for the runnable
        CurrentRunnable.MinimumStartInterval = 0.0
        
        # Set whether the runnable can be invoked concurrently
        CurrentRunnable.CanBeInvokedConcurrently = c
        
        # Set the symbol for the runnable
        CurrentRunnable.Symbol = b
        
        # Create the RTE event associated with the runnable
        createrteevent(CurrentRunnable, d, e, f)
    
    # Create ports for the current software component
    CreatePorts()
    
    # Create shared interfaces for the current software component
    createsharedinterfaces()

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def CreatePorts():
    # Read the port types and names from the specified columns in the ports data
    PortType, PortName = read_columns(ports, 'B', 'C')
    
    # Iterate over the port names and their corresponding types
    for port_name, port_type in zip(PortName, PortType):
        # Check if the port type is 'ReceiverPort'
        if port_type == 'ReceiverPort':
            # Add a new Receiver Port Prototype to the current SWC with the given port name
            CurrentRport = CurrentSWC.Ports.AddNewRPortPrototype(port_name)
        # Check if the port type is 'SenderPort'
        elif port_type == 'SenderPort':
            # Add a new Sender Port Prototype to the current SWC with the given port name
            CurrentPport = CurrentSWC.Ports.AddNewPPortPrototype(port_name)
        else:
            # Print a message if the port type is unknown
            print(f"Unknown port type: {port_type} for port {port_name}")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def createsharedinterfaces():
    # Read the relevant columns from the ports data
    IF_type_col, IF_name_col, DE_col, Argument_col = read_columns(ports, 'D', 'G')
    
    # Initialize a dictionary to hold collections of different interface types
    interface_collections = {
        'SenderReceiverInterface': [],
        'NvDataInterface': [],
        'ParameterInterface': [],
        'ModeSwitchInterface': [],
        'ClientServerInterface': [],
        'TriggerInterface': []
    }
    
    current_interface = None  # To hold the current interface being processed
    current_type = None  # To hold the type of the current interface
    
    # Iterate through the interface names
    for i in range(len(IF_name_col)):
        name = str(IF_name_col[i]).strip()  # Get and clean the interface name
        interface_type = str(IF_type_col[i]).strip()  # Get and clean the interface type
        data_element = str(DE_col[i]).strip()  # Get and clean the data element
        
        # Check if the name is valid (not empty or 'nan')
        if name and name != 'nan':
            # If there is a current interface, add it to the appropriate collection
            if current_interface:
                if current_type in interface_collections:
                    interface_collections[current_type].append(current_interface)
            
            # Create a new interface dictionary for the current name and data element
            current_interface = {
                'name': name,
                'data_elements': [data_element] if data_element and data_element != 'nan' else []
            }
            current_type = interface_type  # Update the current type to the new interface type
        else:
            # If the name is not valid, check if we can add data elements to the current interface
            if current_interface and data_element and data_element != 'nan':
                current_interface['data_elements'].append(data_element)
    
    # After the loop, add the last processed interface to its collection if valid
    if current_interface and current_type:
        if current_type in interface_collections:
            interface_collections[current_type].append(current_interface)
    
    # Map interface types to their corresponding creation functions
    interface_creation_map = {
        'SenderReceiverInterface': SenderReceiverInterface,
        'NvDataInterface': NvDataInterface,
        'ParameterInterface': ParameterInterface,
        'ModeSwitchInterface': ModeSwitchInterface,
        'ClientServerInterface': ClientServerInterface,
        'TriggerInterface': TriggerInterface
    }
    
    # Iterate through each interface type and its collected interfaces
    for interface_type, interfaces in interface_collections.items():
        if interface_type in interface_creation_map:
            creation_func = interface_creation_map[interface_type]  # Get the creation function
            for interface in interfaces:
                try:
                    # Attempt to create the interface using the creation function
                    creation_func(interface['name'], interface['data_elements'])
                except Exception as e:
                    # Print an error message if the creation fails
                    print(f"Error creating {interface_type} {interface['name']}: {e}")

def SenderReceiverInterface(currentIF_name, DataElements):
    # Get the shared package for SenderReceiver interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/SenderReceiver")
    
    # Create a new SenderReceiver interface with the specified name
    currentSRIF = SharedIFpkg.Elements.AddNewSenderReceiverInterface(currentIF_name)
    
    # Set the interface as a non-service interface
    currentSRIF.SetNewIsService().SetValue(False)
    
    # Loop through the provided data elements and add them to the interface
    for itsDE in DataElements:
        currentSRIFDE = currentSRIF.DataElements.AddNew(itsDE)
        
        # Set a description for the data element
        Utilities.SetDescription(currentSRIFDE, "Turn switch sensor position [left, off, right]")
        
        # Set the invalidation policy for the data element
        Utilities.SetInvalidationPolicy(currentSRIF, currentSRIFDE, SystemDeskEnums.HandleInvalidEnum.DontInvalidate)

def NvDataInterface(currentIF_name, NvDatas):
    # Get the shared package for NvData interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/NvData")
    
    # Create a new NvData interface with the specified name
    currentNVIF = SharedIFpkg.Elements.AddNewNvDataInterface(currentIF_name)
    
    # Set the interface as a non-service interface
    currentNVIF.SetNewIsService().SetValue(False)
    
    # Loop through the provided NvData elements and add them to the interface
    for itsDE in NvDatas:
        currentNVIFDE = currentNVIF.NvDatas.AddNew(itsDE)
        
        # Set a description for the NvData element
        Utilities.SetDescription(currentNVIFDE, "Turn switch sensor position [left, off, right]")

def ParameterInterface(currentIF_name, Parameters):
    # Get the shared package for Parameter interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Parameter")
    
    # Create a new Parameter interface with the specified name
    currentPRMIF = SharedIFpkg.Elements.AddNewParameterInterface(currentIF_name)
    
    # Set the interface as a non-service interface
    currentPRMIF.SetNewIsService().SetValue(False)
    
    # Loop through the provided parameter elements and add them to the interface
    for itsDE in Parameters:
        currentPRMIFDE = currentPRMIF.Parameters.AddNew(itsDE)
        
        # Set a description for the parameter element
        Utilities.SetDescription(currentPRMIFDE, "Turn switch sensor position [left, off, right]")

def ModeSwitchInterface(currentIF_name, ModeDeclaration):
    # Get the shared package for ModeSwitch interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ModeSwitch")
    # Functionality for ModeSwitchInterface is not implemented yet

def ClientServerInterface(currentIF_name, Operations):
    # Get the shared package for ClientServer interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/ClientServer")
    
    # Create a new ClientServer interface with the specified name
    currentCSIF = SharedIFpkg.Elements.AddNewClientServerInterface(currentIF_name)
    
    # Set the interface as a non-service interface
    currentCSIF.SetNewIsService().SetValue(False)
    
    # Loop through the provided operations and add them to the interface
    for itsDE in Operations:
        currentCSIFDE = currentCSIF.Operations.AddNew(itsDE)
        
        # Set a description for the operation element
        Utilities.SetDescription(currentCSIFDE, "Turn switch sensor position [left, off, right]") 

def TriggerInterface(currentIF_name, trigger):
    # Get the shared package for Trigger interfaces
    SharedIFpkg = Utilities.GetElementByPath("/SharedElements/PortInterfaces/Trigger")
    # Functionality for TriggerInterface is not implemented yet

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def create_shared_parameter(CurrentInternalBehaviors, b):
    CurrentSharedParameter = CurrentInternalBehaviors.SharedParameters.AddNew(b)
    CurrentSharedParameter.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentSharedParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentSharedParameter, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentSharedParameter, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentSharedParameter)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_per_instance_parameter(CurrentInternalBehaviors, b):
    CurrentPerInstanceParameter = CurrentInternalBehaviors.PerInstanceParameters.AddNew(b)
    CurrentPerInstanceParameter.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentPerInstanceParameter, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentPerInstanceParameter, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentPerInstanceParameter, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentPerInstanceParameter)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_implicit_inter_runnable_variable(CurrentInternalBehaviors, b):
    CurrentImplicitInterRunnableVariables = CurrentInternalBehaviors.ImplicitInterRunnableVariables.AddNew(b)
    CurrentImplicitInterRunnableVariables.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentImplicitInterRunnableVariables, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentImplicitInterRunnableVariables, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentImplicitInterRunnableVariables)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_explicit_inter_runnable_variable(CurrentInternalBehaviors, b):
    CurrentExplicitInterRunnableVariable = CurrentInternalBehaviors.ExplicitInterRunnableVariables.AddNew(b)
    CurrentExplicitInterRunnableVariable.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentExplicitInterRunnableVariable, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentExplicitInterRunnableVariable, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentExplicitInterRunnableVariable)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_per_instance_memory(CurrentInternalBehaviors, b):
    CurrentPerInstanceMemory = CurrentInternalBehaviors.PerInstanceMemorys.AddNew(b) 
    # No additional setup is specified for per-instance memory

def create_ar_typed_per_instance_memory(CurrentInternalBehaviors, b):
    CurrentArTypedPerInstanceMemory = CurrentInternalBehaviors.ArTypedPerInstanceMemorys.AddNew(b)
    CurrentArTypedPerInstanceMemory.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentArTypedPerInstanceMemory, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentArTypedPerInstanceMemory, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentArTypedPerInstanceMemory)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_static_memory(CurrentInternalBehaviors, b):
    CurrentStaticMemorys = CurrentInternalBehaviors.StaticMemorys.AddNew(b)
    CurrentStaticMemorys.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentStaticMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentStaticMemorys, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentStaticMemorys, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentStaticMemorys)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def create_constant_memory(CurrentInternalBehaviors, b):
    CurrentConstantMemorys = CurrentInternalBehaviors.ConstantMemorys.AddNew(b)
    CurrentConstantMemorys.Category = "VALUE"
    Utilities.SetSwImplPolicy(CurrentConstantMemorys, SystemDeskEnums.SwImplPolicyEnum.Standard)
    Utilities.SetSwCalibrationAccess(CurrentConstantMemorys, SystemDeskEnums.SwCalibrationAccessEnum.ReadWrite)
    maxCountAdt = Utilities.GetOrCreateApplicationPrimitiveDataType(CurrentConstantMemorys, "adt_max_count")
    maxCountCm = Utilities.GetOrCreateCompuMethod(maxCountAdt, "cm_max_count")
    Utilities.SetLinearCompuMethod(maxCountCm, 1.0, 0.0)
    maxCountDc = Utilities.GetOrCreateDataConstr(maxCountAdt)
    Utilities.SetPhysicalConstraints(maxCountDc, 10, 200)
    maxCountConst = Utilities.GetOrCreateInitValueConstant(CurrentConstantMemorys)
    Utilities.SetApplicationPrimitiveValueSpecification(maxCountConst, 50)

def CreateOtherIBElements(CurrentInternalBehaviors):
    IBVariableType, IBVariableName, ApplicationDataTypeName, Initvalue, AccessingRunnable, SwCalibrationAccess, SwImplementationPolicy = read_columns(ib_data, 'B', 'H')

    switcher = {
        'SharedParameter': create_shared_parameter,
        'PerInstanceParameter': create_per_instance_parameter,
        'ImplicitInterRunnableVariables': create_implicit_inter_runnable_variable,
        'ExplicitInterRunnableVariable': create_explicit_inter_runnable_variable,
        'PerInstanceMemory': create_per_instance_memory,
        'ArTypedPerInstanceMemory': create_ar_typed_per_instance_memory,
        'StaticMemory': create_static_memory,
        'ConstantMemory': create_constant_memory
    }

    for a, b in zip(IBVariableType, IBVariableName):
        func = switcher.get(a)
        if func:
            func(CurrentInternalBehaviors, b)
        else:
            print("No Other Internal Behaviour Elements present for this software component")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def createrteevent(RunnableName, RTEEventName, RTEEventType, RTEEventInfo):
    # Check the type of RTE event and call the corresponding function
    if RTEEventType == 'AsynchronousServerCallReturnsEvent':
        # Handle asynchronous server call return event
        AsynchronousServerCallReturnsEvent(RTEEventName, RunnableName)
    elif RTEEventType == 'InitEvent':
        # Handle initialization event
        InitEvent(RTEEventName, RunnableName)
    elif RTEEventType == 'BackgroundEvent':
        # Handle background event
        BackgroundEvent(RTEEventName, RunnableName)
    elif RTEEventType == 'TimingEvent':
        # Handle timing event with additional information
        TimingEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'DataReceiveErrorEvent':
        # Handle data receive error event with additional information
        DataReceiveErrorEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'DataSendCompletedEvent':
        # Handle data send completed event with additional information
        DataSendCompletedEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'DataWriteCompletedEvent':
        # Handle data write completed event with additional information
        DataWriteCompletedEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'ExternalTriggerOccurredEvent':
        # Handle external trigger occurred event with additional information
        ExternalTriggerOccurredEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'InternalTriggerOccurredEvent':
        # Handle internal trigger occurred event with additional information
        InternalTriggerOccurredEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'ModeSwitchedAckEvent':
        # Handle mode switched acknowledgment event with additional information
        ModeSwitchedAckEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'OperationInvokedEvent':
        # Handle operation invoked event with additional information
        OperationInvokedEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'SwcModeManagerErrorEvent':
        # Handle software component mode manager error event with additional information
        SwcModeManagerErrorEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'DataReceivedEvent':
        # Handle data received event with additional information
        DataReceivedEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'SwcModeSwitchEvent':
        # Handle software component mode switch event with additional information
        SwcModeSwitchEvent(RTEEventName, RunnableName, RTEEventInfo)
    elif RTEEventType == 'TransformerHardErrorEvent':
        # Handle transformer hard error event with additional information
        TransformerHardErrorEvent(RTEEventName, RunnableName, RTEEventInfo)
    else:
        # Print a message for unrecognized event types
        print(f"Unrecognized event type: {RTEEventName}")

def InitEvent(CurrentRteEvent, AssociatedRunnable):
    # Create a new initialization event and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInitEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable

def TimingEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new timing event with a specified period and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTimingEvent(CurrentRteEvent)
    CurrentEvent.Period = RTEEventInfo
    CurrentEvent.StartOnEventRef = AssociatedRunnable

def AsynchronousServerCallReturnsEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new event for when an asynchronous server call returns
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewAsynchronousServerCallReturnsEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def DataReceiveErrorEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new event for data receive errors and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceiveErrorEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo

def DataReceivedEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new event for when data is received and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataReceivedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.DataIref = RTEEventInfo

def DataSendCompletedEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new event for when data sending is completed and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataSendCompletedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def DataWriteCompletedEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new event for when data writing is completed and associate it with the given runnable
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewDataWriteCompletedEvent(CurrentRteEvent)
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    CurrentEvent.EventSourceRef = RTEEventInfo

def ExternalTriggerOccurredEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new external trigger occurred event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewExternalTriggerOccurredEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the trigger information reference for this event
    CurrentEvent.TriggerIref = RTEEventInfo

def InternalTriggerOccurredEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new internal trigger occurred event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewInternalTriggerOccurredEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the event source reference for this event
    CurrentEvent.EventSourceRef = RTEEventInfo

def ModeSwitchedAckEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new mode switched acknowledgment event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewModeSwitchedAckEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the event source reference for this event
    CurrentEvent.EventSourceRef = RTEEventInfo

def OperationInvokedEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new operation invoked event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewOperationInvokedEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the operation information reference for this event
    CurrentEvent.OperationIref = RTEEventInfo

def SwcModeManagerErrorEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new SWC mode manager error event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeManagerErrorEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the mode group information reference for this event
    CurrentEvent.ModeGroupIref = RTEEventInfo

def SwcModeSwitchEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new SWC mode switch event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewSwcModeSwitchEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the mode information reference for this event
    CurrentEvent.ModeIref = RTEEventInfo
    # Set the activation information for this event
    CurrentEvent.Activation = RTEEventInfo

def TransformerHardErrorEvent(CurrentRteEvent, AssociatedRunnable, RTEEventInfo):
    # Create a new transformer hard error event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewTransformerHardErrorEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable
    # Set the data information reference for this event
    CurrentEvent.DataIref = RTEEventInfo

def BackgroundEvent(CurrentRteEvent, AssociatedRunnable):
    # Create a new background event in the current internal behaviors
    CurrentEvent = CurrentInternalBehaviors.Events.AddNewBackgroundEvent(CurrentRteEvent)
    # Set the reference to the associated runnable for this event
    CurrentEvent.StartOnEventRef = AssociatedRunnable

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def createcompumethod():
    # Read columns from the ADT_Primitive data source
    CompuMethodName, CompuMethodCategory, CompuScaleOROffset, EnumStatesORLSB, Unit = read_columns(adt_primitive, 'D', 'H')
    
    # Get the CompuMethods package and the common package for units
    CompuMethodsPackage = Utilities.GetElementByPath("/SharedElements/CompuMethods")
    commonPackage = Utilities.GetElementByPath("/AUTOSAR/AUTOSAR_PhysicalUnits/Units")
    
    # Initialize variables to hold the current method's details
    current_compu_method_name = None
    current_compu_method_category = None
    current_unit = None
    compu_scale = []  # List to hold scale or offset values
    enum_states = []  # List to hold enumeration states or LSB values
    
    # Iterate through each entry in the CompuMethodName list
    for i in range(len(CompuMethodName)):
        # Check if the current CompuMethodName is not empty
        if CompuMethodName[i]:
            # If we have a previous method, process it
            if current_compu_method_name is not None:
                # Create a new CompuMethod in the package
                currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
                
                # Get the current unit from the common package
                CurrentUnit = commonPackage.Elements.Item(current_unit)
                
                # Define a mapping of categories to their respective handling functions
                switcher = {
                    'IDENTICAL': handle_identical,
                    'TEXTTABLE': handle_texttable,
                    'LINEAR': handle_linear,
                    'SCALE_LINEAR': handle_scale_linear,
                    'SCALE_LINEAR_AND_TEXTTABLE': handle_scale_linear_and_texttable,
                    'RAT_FUNC': handle_rat_func,
                    'SCALE_RAT_FUNC': handle_scale_rat_func,
                    'SCALE_RATIONAL_AND_TEXTTABLE': handle_scale_rational_and_texttable,
                    'TAB_NOINTP': handle_tab_nointp,
                    'BITFIELD_TEXTTABLE': handle_bitfield_texttable,
                }
                
                # Get the appropriate function based on the current category
                func = switcher.get(current_compu_method_category, handle_identical)
                
                # Call the function with the current method's details
                func(currentcompumethod, compu_scale, enum_states, CurrentUnit)
            
            # Update the current method's details for the next iteration
            current_compu_method_name = CompuMethodName[i]
            current_compu_method_category = CompuMethodCategory[i]
            current_unit = Unit[i]
            compu_scale = []  # Reset the scale list for the new method
            enum_states = []  # Reset the enum states list for the new method
        
        # If there are scale or offset values, add them to the list
        if i < len(CompuScaleOROffset) and CompuScaleOROffset[i] is not None:
            compu_scale.append(CompuScaleOROffset[i])
        
        # If there are enumeration states or LSB values, add them to the list
        if i < len(EnumStatesORLSB) and EnumStatesORLSB[i] is not None:
            enum_states.append(EnumStatesORLSB[i])
    
    # After the loop, process the last CompuMethod if it exists
    if current_compu_method_name is not None:
        currentcompumethod = CompuMethodsPackage.Elements.AddNewCompuMethod(current_compu_method_name)
        CurrentUnit = commonPackage.Elements.Item(current_unit)
        
        # Get the appropriate function based on the current category
        switcher = {
            'IDENTICAL': handle_identical,
            'TEXTTABLE': handle_texttable,
            'LINEAR': handle_linear,
            'SCALE_LINEAR': handle_scale_linear,
            'SCALE_LINEAR_AND_TEXTTABLE': handle_scale_linear_and_texttable,
            'RAT_FUNC': handle_rat_func,
            'SCALE_RAT_FUNC': handle_scale_rat_func,
            'SCALE_RATIONAL_AND_TEXTTABLE': handle_scale_rational_and_texttable,
            'TAB_NOINTP': handle_tab_nointp,
            'BITFIELD_TEXTTABLE': handle_bitfield_texttable,
        }
        
        # Get the appropriate function and call it
        func = switcher.get(current_compu_method_category, handle_identical)
        func(currentcompumethod, compu_scale, enum_states, CurrentUnit)

def handle_identical(currentcompumethod, NA, NA1, Current_Unit):
    # Check if NA is None or an empty string
    if NA is None or NA == "":
        print("CompuMethodInfo is not applicable for IDENTICAL category.")
    # Check if NA1 is None or an empty string
    if NA1 is None or NA1 == "":
        print("CompuMethodInfo is not applicable for IDENTICAL category.")
    # Check if Current_Unit is None or an empty string
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for IDENTICAL category.")
    
    # Set the category and unit reference for the current computation method
    currentcompumethod.Category = 'IDENTICAL'
    currentcompumethod.UnitRef = Current_Unit

def handle_texttable(currentcompumethod, Compu_Scale, Enum_States, Current_Unit):
    defaultValue = None  # Initialize default value as None
    
    # Check if Compu_Scale is None or an empty string
    if Compu_Scale is None or Compu_Scale == "":
        print("CompuMethodInfo is not applicable for TEXTTABLE category.")
    # Check if Enum_States is None or an empty string
    if Enum_States is None or Enum_States == "":
        print("CompuMethodInfo is not applicable for TEXTTABLE category.")
    # Check if Current_Unit is None or an empty string
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for TEXTTABLE category.")
    
    # Set the category and display format for the current computation method
    currentcompumethod.Category = 'TEXTTABLE'
    currentcompumethod.DisplayFormat = r"%g"
    currentcompumethod.UnitRef = Current_Unit
    
    # Create a new internal to physical mapping for the computation method
    compuInternalToPhys = currentcompumethod.SetNewCompuInternalToPhys()
    
    # Loop through Compu_Scale and Enum_States to create compu scales
    for a, b in zip(Compu_Scale, Enum_States):
        lowerLimit = a  # Set lower limit
        upperLimit = a  # Set upper limit
        symbol = b      # Set symbol
        
        # Create a new compu scale
        compuScale = compuInternalToPhys.CompuScales.AddNew()
        compuScale.Symbol = symbol
        
        # Set lower limit for the compu scale
        compuScaleLL = compuScale.SetNewLowerLimit()
        compuScaleLL.MixedContent.AddStringContent(str(lowerLimit))
        compuScaleLL.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        
        # Set upper limit for the compu scale
        compuScaleUL = compuScale.SetNewUpperLimit()
        compuScaleUL.MixedContent.AddStringContent(str(upperLimit))
        compuScaleUL.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        
        # Set short label and constant for the compu scale
        compuScale.ShortLabel = symbol
        compuConst = compuScale.SetNewCompuConst()
        compuConst.Vt = symbol
    
    # If defaultValue is set, print it (currently always None)
    if defaultValue is not None:
        print(defaultValue)
        compuInternalToPhys.CompuDefaultValue.Vt = defaultValue

def handle_linear(currentcompumethod, CompuScaleOROffset, EnumStatesORLSB, Current_Unit):
    # Check if CompuScaleOROffset is None or an empty string
    if CompuScaleOROffset is None or CompuScaleOROffset == "":
        print("CompuMethodInfo is not applicable for LINEAR category.")
    # Check if Current_Unit is None or an empty string
    if Current_Unit is None or Current_Unit == "":
        print("Unit is not applicable for LINEAR category.")

def handle_scale_linear(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_LINEAR category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_LINEAR category.")

def handle_scale_linear_and_texttable(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_LINEAR_AND_TEXTTABLE category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_LINEAR_AND_TEXTTABLE category.")

def handle_rat_func(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for RAT_FUNC category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for RAT_FUNC category.")

def handle_scale_rat_func(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_RAT_FUNC category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_RAT_FUNC category.")

def handle_scale_rational_and_texttable(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for SCALE_RATIONAL_AND_TEXTTABLE category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for SCALE_RATIONAL_AND_TEXTTABLE category.")

def handle_tab_nointp(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for TAB_NOINTP category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for TAB_NOINTP category.")

def handle_bitfield_texttable(compu_method_info, unit):
    # Check if compu_method_info is None or an empty string
    if compu_method_info is None or compu_method_info == "":
        print("CompuMethodInfo is not applicable for BITFIELD_TEXTTABLE category.")
    # Check if unit is None or an empty string
    if unit is None or unit == "":
        print("Unit is not applicable for BITFIELD_TEXTTABLE category.")

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def createDC():
    # Read columns from the adt_primitive data source
    DataConstraintName, DataConstraintType, Min, Max = read_columns(adt_primitive, 'I', 'L')
    
    # Get the Data Constraints package from the shared elements
    DCPackage = Utilities.GetElementByPath("/SharedElements/DataConstr")
    
    # Iterate through the data constraints
    for a, b, c, d in zip(DataConstraintName, DataConstraintType, Min, Max):
        constrLevel = 0  # Initialize constraint level
        lowerLimit = c   # Set lower limit from the read data
        upperLimit = d   # Set upper limit from the read data
        
        # Check if the constraint type is 'internalConstrs'
        if b == 'internalConstrs':
            # Create a new internal data constraint
            dataConstr = DCPackage.Elements.AddNewDataConstr(a)
            dataConstrRule = dataConstr.DataConstrRules.AddNew()
            
            # Set the constraint level if it's not None
            if constrLevel is not None:
                dataConstrRule.ConstrLevel = constrLevel
            
            # Set new internal constraints
            internalConstrs = dataConstrRule.SetNewInternalConstrs()
            
            # Set the lower limit if it is not None
            if lowerLimit is not None:
                lowLimit = internalConstrs.SetNewLowerLimit()
                lowLimit.MixedContent.AddStringContent(Value2Str(lowerLimit))
                lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
            
            # Set the upper limit if it is not None
            if upperLimit is not None:
                upLimit = internalConstrs.SetNewUpperLimit()
                upLimit.MixedContent.AddStringContent(Value2Str(upperLimit))
                upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        
        # Check if the constraint type is 'physConstrs'
        elif b == 'physConstrs':
            # Create a new physical data constraint
            dataConstr = DCPackage.Elements.AddNewDataConstr(a)
            dataConstrRule = dataConstr.DataConstrRules.AddNew()
            
            # Set the constraint level if it's not None
            if constrLevel is not None:
                dataConstrRule.ConstrLevel = constrLevel
            
            # Set new physical constraints
            physConstrs = dataConstrRule.SetNewPhysConstrs()
            
            # Set the lower limit if it is not None
            if lowerLimit is not None:
                lowLimit = physConstrs.SetNewLowerLimit()
                lowLimit.MixedContent.AddStringContent(Value2Str(lowerLimit))
                lowLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
            
            # Set the upper limit if it is not None
            if upperLimit is not None:
                upLimit = physConstrs.SetNewUpperLimit()
                upLimit.MixedContent.AddStringContent(Value2Str(upperLimit))
                upLimit.IntervalType = SystemDeskEnums.IntervalTypeEnum.Closed
        
        # If the constraint type is neither 'internalConstrs' nor 'physConstrs'
        else:
            if b is not None:
                print(f"Value of b: {b}")  # Print the value of b for debugging purposes

#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#
#               # ####################### =============== %%%%%%%%% ++++++++++++ ---------- __________ SECTION :  abcd __________ ----------  ++++++++++++ %%%%%%%%% =============== ####################### #               #
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------#

def Main():
    # Execute the main sequence of functions for project setup
    read_user_defined_excel()  # Read user-defined data from an Excel file
    ProjectConfig()         # Configure the project settings
    CreateProjectAndPackages()  # Create the project and its packages
    CleanupProjectAndPackages()  # Clean up any temporary project data
    CreateSwcs()           # Create software components
    createcompumethod()    # Create computation methods
    createDC()             # Create data constraints

# Entry point of the script
if __name__ == "__main__":
    Main()  # Run the main function
    print("Ready")  # Indicate that the process is complete     