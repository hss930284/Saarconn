import xml.etree.ElementTree as ET
import pandas as pd # type: ignore
import os
import openpyxl # type: ignore
from openpyxl.styles import PatternFill # type: ignore

def parse_arxml(arxml_file):
    # Parse the .arxml file using ElementTree
    tree = ET.parse(arxml_file)
    root = tree.getroot()

    # Lists to store Type, Short Name, and UUID data
    types = []
    short_names = []
    uuids = []

    # Iterate over all elements and collect necessary information
    def extract_data(element):
        # If the element has a SHORT-NAME tag, we collect the data
        for child in element:
            if child.tag.endswith('SHORT-NAME'):  # Handle namespace
                # Get the short name text
                short_name = child.text.strip() if child.text else "No Name"
                short_names.append(short_name)
                
                # Get the parent type (the element tag of the parent node)
                parent_type = element.tag.split('}')[1] if '}' in element.tag else element.tag  # Clean namespace
                types.append(parent_type)
                
                # Check if UUID is present for this element
                uuid = element.get('UUID', "No UUID")  # Get UUID if exists
                uuids.append(uuid)

            # Recurse into the child elements
            extract_data(child)

    # Start the extraction from the root element
    extract_data(root)

    # Create a DataFrame to store the results with columns: Type, Short Name, UUID
    data = {
        "Type": types,
        "Short Name": short_names,
        "UUID": uuids
    }
    df = pd.DataFrame(data)

    # Sort the DataFrame by the "Type" column
    df_sorted = df.sort_values(by="Type").reset_index(drop=True)
    
    return df_sorted

# def parse_arxml_from_folder(folder_path, excel_file):
#     # List to hold all ARXML file paths in the specified folder
#     arxml_files = []

#     # Get all .arxml files in the provided folder
#     for filename in os.listdir(folder_path):
#         if filename.lower().endswith('.arxml'):
#             arxml_files.append(os.path.join(folder_path, filename))

#     if not arxml_files:
#         print(f"No ARXML files found in the folder: {folder_path}")
#         return

#     # Create a Pandas Excel writer to save multiple sheets
#     with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
#         # Dictionary to hold DataFrames for each ARXML file
#         dfs = {}
        
#         for arxml_file in arxml_files:
#             # Extract the filename without extension to use as the sheet name
#             sheet_name = os.path.splitext(os.path.basename(arxml_file))[0]
            
#             # Parse the ARXML file and get the sorted DataFrame
#             df_sorted = parse_arxml(arxml_file)
            
#             # Store DataFrame for later comparison
#             dfs[sheet_name] = df_sorted
            
#             # Write the DataFrame to the sheet named after the ARXML file
#             df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)

#         # Highlight common cells across all files
#         highlight_common_cells(excel_file, dfs)

#     print(f"Excel file saved at {excel_file}")

def parse_arxml_from_folder(folder_path, excel_file):
    # Ensure the file extension is .xlsx
    if not excel_file.lower().endswith('.xlsx'):
        print(f"Error: The file '{excel_file}' does not have the correct '.xlsx' extension.")
        return

    # List to hold all ARXML file paths in the specified folder
    arxml_files = []

    # Get all .arxml files in the provided folder
    for filename in os.listdir(folder_path):
        if filename.lower().endswith('.arxml'):
            arxml_files.append(os.path.join(folder_path, filename))

    if not arxml_files:
        print(f"No ARXML files found in the folder: {folder_path}")
        return

    # Create a Pandas Excel writer to save multiple sheets
    with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
        # Dictionary to hold DataFrames for each ARXML file
        dfs = {}
        
        for arxml_file in arxml_files:
            # Extract the filename without extension to use as the sheet name
            sheet_name = os.path.splitext(os.path.basename(arxml_file))[0]
            
            # Parse the ARXML file and get the sorted DataFrame
            df_sorted = parse_arxml(arxml_file)
            
            # Store DataFrame for later comparison
            dfs[sheet_name] = df_sorted
            
            # Write the DataFrame to the sheet named after the ARXML file
            df_sorted.to_excel(writer, sheet_name=sheet_name, index=False)

        # Highlight common cells across all files
        # highlight_common_cells(excel_file, dfs)

    print(f"Excel file saved at {excel_file}")


# def highlight_common_cells(excel_file, dfs):
#     # Load the Excel file with openpyxl to apply formatting

#     print(f"Opening the Excel file: {excel_file}")  # Debugging line to check the file path


#     wb = openpyxl.load_workbook(excel_file)
    
#     # Get all short names and UUIDs that are common across all ARXML files
#     common_data = set.intersection(*[set(zip(df['Short Name'], df['UUID'])) for df in dfs.values()])
    
#     # Loop through each sheet and highlight common cells
#     for sheet_name, df in dfs.items():
#         sheet = wb[sheet_name]
        
#         # Define a yellow fill for highlighting
#         yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
#         # Iterate through rows and highlight common short name and UUID pairs
#         for idx, (short_name, uuid) in enumerate(zip(df['Short Name'], df['UUID']), start=2):  # Starting from row 2
#             if (short_name, uuid) in common_data:
#                 # Highlight the corresponding cells in the "Short Name" and "UUID" columns
#                 sheet[f"B{idx}"].fill = yellow_fill
#                 sheet[f"C{idx}"].fill = yellow_fill

#     # Save the Excel file after applying the highlighting
#     wb.save(excel_file)
#     print("Common cells highlighted.")

# Prompt the user to provide the folder path containing ARXML files and Excel file path
folder_path = input("Please provide the folder path containing the ARXML files: ")
excel_file = input("Please provide the full path to save the Excel file (e.g., output.xlsx): ")

# Run the function
parse_arxml_from_folder(folder_path, excel_file)
